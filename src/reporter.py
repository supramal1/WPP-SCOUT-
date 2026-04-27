import os
import pandas as pd
import numpy as np
import re
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


# =============================================================================
# COLUMN LISTS - Single source of truth
# =============================================================================

# Core columns for Looker Studio export
LOOKER_COLS = [
    "creative_name",
    "platform",
    "currency",
    "format_canonical",
    "placement_canonical",
    "objective_normalized",
    "asset_type_canonical",
    "asset_type_subtype",
    "os_target",
    "device_type",
    "audience_segment",
    "buying_type",
    "campaign_normalized",
    "cross_platform",
    "platforms_active",
    "spend",
    "reach",
    "impressions",
    "frequency",
    "vtr_2s",
    "completion_rate",
    "completion_vs_expected",
    "duration_s",
    "ctr",
    "engagement_rate",
    "share_rate",
    "cost_per_complete_view",
    "attention_proxy_score",
    "attention_inputs_available",
    "score_renormalized",
    "composite_score",
    "tier",
    "action",
    "scoring_group",
    "low_confidence",
]

# Full display columns for detail tabs
DISPLAY_COLS = [
    "creative_name",
    "platform",
    "buying_type",
    "currency",
    "format_canonical",
    "placement_canonical",
    "objective_normalized",
    "scoring_group",
    "composite_score",
    "tier",
    "action",
    "rank_in_group",
    "group_size",
    "n_variants",
    "n_campaigns",
    "spend",
    "reach",
    "impressions",
    "frequency",
    "vtr_2s",
    "completion_rate",
    "completion_vs_expected",
    "duration_s",
    "ctr",
    "engagement_rate",
    "share_rate",
    "cost_per_complete_view",
    "video_views_100",
    "attention_proxy_score",
    "attention_inputs_available",
    "score_renormalized",
    "audience_consistency",
    "low_confidence",
    "freq_penalty",
    "format_flag",
    "explanation",
]

# Summary columns for dashboard tables
DASHBOARD_COLS = [
    "creative_name",
    "platform",
    "buying_type",
    "format_canonical",
    "placement_canonical",
    "objective_normalized",
    "composite_score",
    "tier",
    "action",
    "spend",
    "reach",
    "frequency",
    "vtr_2s",
    "completion_rate",
    "ctr",
    "engagement_rate",
]

# KPI columns available for selection
AVAILABLE_KPI_COLUMNS = [
    ("VTR (2s/3s)", "vtr_2s"),
    ("Hook Rate", "canonical_hook_rate"),
    ("Hold Rate", "canonical_hold_rate"),
    ("Completion Rate", "completion_rate"),
    ("Engagement Rate", "engagement_rate"),
    ("CTR", "ctr"),
    ("CPC", "cost_per_click"),
    ("Reach", "reach"),
    ("Frequency", "frequency"),
    ("Spend", "spend"),
    ("Impressions", "impressions"),
]

XL_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

ET.register_namespace("", XL_MAIN_NS)
ET.register_namespace("r", DOC_REL_NS)


# =============================================================================
# CONSOLE REPORT
# =============================================================================


def print_console_report(df: pd.DataFrame, brand: str = ""):
    """Print a summary to the console."""
    title = f"CREATIVE PERFORMANCE ANALYSIS{f' — {brand.upper()}' if brand else ''}"
    print("\n" + "=" * 80)
    print(title)
    print("=" * 80)

    # Overview
    total = len(df)
    high_conf = len(df[~df["low_confidence"]])
    print(f"\nTotal creatives analyzed: {total}")
    print(f"Statistically significant (>€500 spend & >10k reach): {high_conf}")
    print(f"Low confidence: {total - high_conf}")

    for platform, buying_type in [
        ("TikTok", "Paid"),
        ("Meta", "Paid"),
        ("TikTok", "Boosting"),
        ("Meta", "Boosting"),
    ]:
        platform_df = df[
            (df["platform"] == platform) & (df["buying_type"] == buying_type)
        ]
        if platform_df.empty:
            continue

        print(f"\n{'─' * 80}")
        print(f"  {platform.upper()} ({buying_type.upper()}) - TOP PERFORMERS")
        print(f"{'─' * 80}")
        print(
            f"  NOTE: Scores are only comparable within this cohort (same platform + buying type)"
        )

        # Show top 10 high-confidence creatives
        top = platform_df[~platform_df["low_confidence"]].head(10)
        for i, (_, row) in enumerate(top.iterrows(), 1):
            score = row["composite_score"]
            tier_icon = "★" if score >= 70 else "●"
            print(f"\n  {tier_icon} #{i} | Score: {score}/100 | {row['tier']}")
            print(f"    Creative: {row['creative_name'][:60]}")
            print(
                f"    Objective: {row.get('objective_normalized', row.get('objective', 'Unknown'))} | Format: {row.get('format_canonical', row.get('format', 'Unknown'))}"
            )
            print(
                f"    Spend: €{row['spend']:,.0f} | Reach: {row['reach']:,.0f} | Freq: {row['frequency']:.1f}x"
            )
            if row["vtr_2s"] > 0:
                print(
                    f"    2s VTR: {row['vtr_2s']:.1f}% | Completion: {row['completion_rate']:.2f}%"
                )
            # Show attention proxy if available
            if row.get("attention_inputs_available", False):
                print(
                    f"    Attention Proxy: {row.get('attention_proxy_score', 0):.1f}/100"
                )
            if row.get("score_renormalized", False):
                print(f"    (Score renormalized - no attention inputs)")
            if row.get("format_flag"):
                print(f"    ** {row['format_flag']}")

        # Bottom 5
        print(f"\n  {'─' * 40}")
        print(f"  {platform.upper()} - UNDERPERFORMERS (consider pausing)")
        bottom = platform_df[~platform_df["low_confidence"]].tail(5)
        for _, row in bottom.iterrows():
            print(
                f"    Score: {row['composite_score']}/100 | {row['creative_name'][:50]} | Freq: {row['frequency']:.1f}x"
            )

    print(f"\n{'=' * 80}")
    print("Full results exported to Excel.")
    print("=" * 80 + "\n")


# =============================================================================
# EXCEL EXPORT
# =============================================================================


def export_excel(
    df: pd.DataFrame,
    output_path: str,
    df_raw: "pd.DataFrame | None" = None,
    brand: str = "",
    target: str = "excel",
):
    """Export scored results to a formatted workbook with a formula-driven interactive dashboard."""
    # Ensure all display columns exist
    for col in DISPLAY_COLS:
        if col not in df.columns:
            df[col] = ""

    paid = df[df["buying_type"] == "Paid"] if "buying_type" in df.columns else df
    boosting = (
        df[df["buying_type"] == "Boosting"]
        if "buying_type" in df.columns
        else pd.DataFrame()
    )
    total = len(paid)
    high_conf = len(paid[~paid["low_confidence"]])
    tiktok_count = len(paid[paid["platform"] == "TikTok"])
    meta_count = len(paid[paid["platform"] == "Meta"])

    # Prepare all data for helper sheet (combine paid + boosting for "Both" option)
    all_data = (
        pd.concat([paid, boosting], ignore_index=True)
        if not boosting.empty
        else paid.copy()
    )

    output_dir = os.path.dirname(output_path) or "."
    os.makedirs(output_dir, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        prefix=".creative-analyzer-",
        suffix=".xlsx",
        dir=output_dir,
        delete=False,
    ) as tmp_file:
        temp_output_path = tmp_file.name

    try:
        writer = pd.ExcelWriter(temp_output_path, engine="xlsxwriter")
        workbook = writer.book

        # --- Shared formats ---
        header_fmt = workbook.add_format(
            {
                "bold": True,
                "bg_color": "#1a1a2e",
                "font_color": "white",
                "border": 1,
                "text_wrap": True,
                "valign": "top",
            }
        )
        wrap_fmt = workbook.add_format({"text_wrap": True, "valign": "top"})

        # Dashboard formats
        title_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 20,
                "font_color": "#1a1a2e",
                "bottom": 2,
                "bottom_color": "#1a1a2e",
            }
        )
        subtitle_fmt = workbook.add_format(
            {
                "font_size": 12,
                "font_color": "#555555",
                "italic": True,
            }
        )
        section_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 13,
                "font_color": "#1a1a2e",
                "bottom": 1,
                "bottom_color": "#cccccc",
            }
        )
        filter_label_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 10,
                "font_color": "#555555",
            }
        )
        filter_value_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 11,
                "bg_color": "#e8f5e9",
                "border": 1,
                "border_color": "#c8e6c9",
            }
        )
        body_fmt = workbook.add_format(
            {
                "font_size": 11,
                "text_wrap": True,
                "valign": "top",
            }
        )
        bold_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 11,
                "text_wrap": True,
                "valign": "top",
            }
        )
        score_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 11,
                "align": "center",
                "border": 1,
                "border_color": "#e0e0e0",
                "num_format": "0.0",
            }
        )

        # Tier formats
        tier_top_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 11,
                "bg_color": "#d4edda",
                "border": 1,
                "border_color": "#c3e6cb",
            }
        )
        tier_strong_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 11,
                "bg_color": "#cce5ff",
                "border": 1,
                "border_color": "#b8daff",
            }
        )
        tier_avg_fmt = workbook.add_format(
            {
                "font_size": 11,
                "bg_color": "#fff3cd",
                "border": 1,
                "border_color": "#ffeeba",
            }
        )
        tier_below_fmt = workbook.add_format(
            {
                "font_size": 11,
                "bg_color": "#fde2e2",
                "border": 1,
                "border_color": "#f5c6cb",
            }
        )

        # =====================================================================
        # SHEET 0: DASHBOARD (Formula-driven with Controls)
        # Create Dashboard FIRST so it appears as the first sheet
        # =====================================================================
        ws = workbook.add_worksheet("Dashboard")
        ws.hide_gridlines(2)
        ws.set_tab_color("#1a1a2e")

        # Column widths for Dashboard
        ws.set_column("A:A", 3)  # margin
        ws.set_column("B:B", 18)  # rank/labels
        ws.set_column("C:C", 45)  # creative
        ws.set_column("D:D", 10)  # score
        ws.set_column("E:H", 12)  # KPIs
        ws.set_column("I:I", 55)  # explanation
        ws.set_column("J:J", 3)  # margin

        # --- TITLE ---
        row = 1
        dash_title = f"WPP Scout Dashboard{f' — {brand}' if brand else ''}"
        ws.merge_range(row, 1, row, 8, dash_title, title_fmt)
        row += 1
        ws.merge_range(
            row,
            1,
            row,
            8,
            "✓ Dropdowns are now ACTIVE. Change any selector to filter data dynamically. Requires Excel 2021/365.",
            subtitle_fmt,
        )
        row += 2

        # --- CONTROL PANEL ---
        ws.merge_range(row, 1, row, 8, "FILTER CONTROLS (Active)", section_fmt)
        row += 1

        # Prepare unique values for dropdowns
        campaigns = (
            ["All"]
            + sorted(
                [
                    str(x)
                    for x in all_data["campaign_normalized"].dropna().unique()
                    if str(x).strip() and str(x) != "nan"
                ]
            )
            if "campaign_normalized" in all_data.columns
            else ["All"]
        )
        platforms = ["All", "TikTok", "Meta"]
        formats = ["All", "Video", "Motion", "Static"]
        top_ranges = ["5", "10", "25", "50"]
        dimensions = ["Asset Type", "Placement", "OS", "Device", "Objective"]
        kpi_options = [
            "VTR",
            "Completion Rate",
            "CTR",
            "Engagement Rate",
            "Spend",
            "Reach",
        ]

        # Control row 1: Campaign, Platform, Format
        control_row1 = row
        ws.write(control_row1, 1, "CAMPAIGN:", filter_label_fmt)
        ws.write(control_row1, 2, "All", filter_value_fmt)
        ws.write(control_row1, 3, "PLATFORM:", filter_label_fmt)
        ws.write(control_row1, 4, "All", filter_value_fmt)
        ws.write(control_row1, 5, "FORMAT:", filter_label_fmt)
        ws.write(control_row1, 6, "All", filter_value_fmt)
        row += 1

        # Add data validation for Campaign
        ws.data_validation(
            control_row1,
            2,
            control_row1,
            2,
            {
                "validate": "list",
                "source": campaigns[:100],  # Excel limit
            },
        )
        ws.data_validation(
            control_row1,
            4,
            control_row1,
            4,
            {
                "validate": "list",
                "source": platforms,
            },
        )
        ws.data_validation(
            control_row1,
            6,
            control_row1,
            6,
            {
                "validate": "list",
                "source": formats,
            },
        )

        # Control row 2: Top Range, Dimension, Buying Type
        control_row2 = row
        ws.write(control_row2, 1, "TOP RANGE:", filter_label_fmt)
        ws.write(control_row2, 2, "10", filter_value_fmt)
        ws.write(control_row2, 3, "DIMENSION:", filter_label_fmt)
        ws.write(control_row2, 4, "Asset Type", filter_value_fmt)
        ws.write(control_row2, 5, "BUYING TYPE:", filter_label_fmt)
        ws.write(control_row2, 6, "Paid", filter_value_fmt)
        row += 1

        ws.data_validation(
            control_row2,
            2,
            control_row2,
            2,
            {
                "validate": "list",
                "source": top_ranges,
            },
        )
        ws.data_validation(
            control_row2,
            4,
            control_row2,
            4,
            {
                "validate": "list",
                "source": dimensions,
            },
        )
        ws.data_validation(
            control_row2,
            6,
            control_row2,
            6,
            {
                "validate": "list",
                "source": ["Paid", "Boosting", "Both"],
            },
        )

        # Control row 3: KPI Slots
        control_row3 = row
        ws.write(control_row3, 1, "KPI SLOT 1:", filter_label_fmt)
        ws.write(control_row3, 2, "VTR", filter_value_fmt)
        ws.write(control_row3, 3, "KPI SLOT 2:", filter_label_fmt)
        ws.write(control_row3, 4, "Completion Rate", filter_value_fmt)
        ws.write(control_row3, 5, "KPI SLOT 3:", filter_label_fmt)
        ws.write(control_row3, 6, "CTR", filter_value_fmt)
        row += 1

        ws.data_validation(
            control_row3,
            2,
            control_row3,
            2,
            {
                "validate": "list",
                "source": kpi_options,
            },
        )
        ws.data_validation(
            control_row3,
            4,
            control_row3,
            4,
            {
                "validate": "list",
                "source": kpi_options,
            },
        )
        ws.data_validation(
            control_row3,
            6,
            control_row3,
            6,
            {
                "validate": "list",
                "source": kpi_options,
            },
        )
        row += 2

        # --- FORMULA-DRIVEN STATS (based on filtered data) ---
        stat_box_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 18,
                "font_color": "#1a1a2e",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
                "border_color": "#e0e0e0",
            }
        )
        stat_label_fmt = workbook.add_format(
            {
                "font_size": 9,
                "font_color": "#888888",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
                "border_color": "#e0e0e0",
            }
        )

        # Dynamic stats using FILTER formulas
        # Filter formula: FILTER(_DataScore, (platform_cond) * (format_cond) * (buying_cond) * (campaign_cond))
        # Cell references: C2=Campaign, E2=Platform, G2=Format, G3=BuyingType
        # Platform condition: IF(E2="All", 1, _DataPlatform=E2)
        filter_formula = (
            "LET("
            "  p_sel, Dashboard!$E$%d, "
            "  f_sel, Dashboard!$G$%d, "
            "  b_sel, Dashboard!$G$%d, "
            "  c_sel, Dashboard!$C$%d, "
            "  filtered, FILTER(_DataScore, "
            '    (IF(p_sel="All", 1, _DataPlatform=p_sel)) * '
            '    (IF(f_sel="All", 1, _DataFormat=f_sel)) * '
            '    (IF(b_sel="Both", 1, _DataBuyingType=b_sel)) * '
            '    (IF(c_sel="All", 1, _DataCampaign=c_sel)) * '
            "    (_DataLowConf=FALSE)), "
            "  COUNT(filtered))"
        ) % (control_row1 + 1, control_row1 + 1, control_row2 + 1, control_row1 + 1)

        # Write dynamic stat boxes with formulas
        stats_row = row
        ws.write_formula(stats_row, 1, filter_formula, stat_box_fmt)
        ws.write_formula(stats_row + 1, 1, "Filtered Count", stat_label_fmt)

        # Total (static reference)
        ws.write(stats_row, 3, str(total), stat_box_fmt)
        ws.write(stats_row + 1, 3, "Total Creatives", stat_label_fmt)

        # TikTok count (formula)
        tiktok_formula = f'=COUNTIF(_DataPlatform,"TikTok")'
        ws.write_formula(stats_row, 5, tiktok_formula, stat_box_fmt)
        ws.write(stats_row + 1, 5, "TikTok", stat_label_fmt)

        # Meta count (formula)
        meta_formula = f'=COUNTIF(_DataPlatform,"Meta")'
        ws.write_formula(stats_row, 7, meta_formula, stat_box_fmt)
        ws.write(stats_row + 1, 7, "Meta", stat_label_fmt)

        row += 3

        # --- TOP PERFORMERS (Formula-Driven) ---
        # Uses Excel FILTER + SORT + TAKE to dynamically show top N based on selectors
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  TOP PERFORMERS (Dynamic)",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#2c3e50",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        # Headers for Top Performers table
        top_headers = [
            "Rank",
            "Creative",
            "Score",
            "Objective / Format",
            "ER",
            "CTR",
            "VTR",
            "Spend",
        ]
        for ci, h in enumerate(top_headers):
            ws.write(
                row,
                ci + 1,
                h,
                workbook.add_format(
                    {
                        "bold": True,
                        "font_size": 9,
                        "font_color": "#888888",
                        "bottom": 1,
                        "bottom_color": "#dddddd",
                    }
                ),
            )
        row += 1

        # Write formula-driven Top 10 (using FILTER + SORT + TAKE)
        # Formula: TAKE(SORT(FILTER(_DataCreative:Score, conditions), score_col, -1), 10)
        # We write placeholder rows with formulas that will populate when opened in Excel 365
        top_formula_base = (
            "=LET("
            "  p_sel, Dashboard!$E$%d, "
            "  f_sel, Dashboard!$G$%d, "
            "  b_sel, Dashboard!$G$%d, "
            "  c_sel, Dashboard!$C$%d, "
            "  n_rows, VALUE(Dashboard!$C$%d), "
            "  filtered, FILTER(HSTACK(_DataCreative, _DataScore, _DataVTR, _DataCompletion, _DataCTR, _DataSpend), "
            '    (IF(p_sel="All", 1, _DataPlatform=p_sel)) * '
            '    (IF(f_sel="All", 1, _DataFormat=f_sel)) * '
            '    (IF(b_sel="Both", 1, _DataBuyingType=b_sel)) * '
            '    (IF(c_sel="All", 1, _DataCampaign=c_sel)) * '
            "    (_DataLowConf=FALSE)), "
            "  sorted, SORT(filtered, 2, -1), "
            "  top_n, TAKE(sorted, n_rows), "
            "  INDEX(top_n, %d, %d))"
        ) % (
            control_row1 + 1,
            control_row1 + 1,
            control_row2 + 1,
            control_row1 + 1,
            control_row2 + 1,
            1,
            1,
        )

        # Pre-rendered fallback (for when formulas don't work in older Excel)
        # Write actual data rows below the formula section
        row += 1
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  TOP 10 (Current Filters - see Rankings tabs for full data)",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 11,
                    "font_color": "#555555",
                    "bg_color": "#f8f9fa",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        # Write actual top 10 from paid data
        top_10 = paid[~paid["low_confidence"]].head(10)
        for rank, (_, r) in enumerate(top_10.iterrows(), 1):
            ws.write(row, 1, rank, workbook.add_format({"align": "center"}))
            ws.write(row, 2, str(r["creative_name"])[:38], body_fmt)
            score_val = r["composite_score"]
            tier = r.get("tier", "Average")
            if tier == "Top Performer":
                ws.write(row, 3, score_val, tier_top_fmt)
            elif tier == "Strong":
                ws.write(row, 3, score_val, tier_strong_fmt)
            elif tier == "Average":
                ws.write(row, 3, score_val, tier_avg_fmt)
            else:
                ws.write(row, 3, score_val, tier_below_fmt)
            # Objective / Format context column
            obj = r.get("objective_normalized", r.get("objective", ""))
            fmt = r.get("format_canonical", r.get("format", ""))
            ws.write(row, 4, f"{obj} · {fmt}", body_fmt)
            ws.write(row, 5, f"{r.get('engagement_rate', 0):.3f}%", body_fmt)
            ws.write(row, 6, f"{r.get('ctr', 0):.2f}%", body_fmt)
            ws.write(row, 7, f"{r.get('vtr_2s', 0):.1f}%", body_fmt)
            ws.write(row, 8, f"€{r.get('spend', 0):,.0f}", body_fmt)
            row += 1

        # --- BOTTOM PERFORMERS ---
        row += 2
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  BOTTOM PERFORMERS (Underperformers to Review)",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#c0392b",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        # Write actual bottom 10 from paid data (high confidence only)
        bottom_10 = (
            paid[~paid["low_confidence"]].tail(10).iloc[::-1]
        )  # Reverse to show worst first
        for rank, (_, r) in enumerate(bottom_10.iterrows(), 1):
            ws.write(row, 1, rank, workbook.add_format({"align": "center"}))
            ws.write(row, 2, str(r["creative_name"])[:38], body_fmt)
            score_val = r["composite_score"]
            ws.write(row, 3, score_val, tier_below_fmt)
            obj = r.get("objective_normalized", r.get("objective", ""))
            fmt = r.get("format_canonical", r.get("format", ""))
            ws.write(row, 4, f"{obj} · {fmt}", body_fmt)
            ws.write(row, 5, f"{r.get('engagement_rate', 0):.3f}%", body_fmt)
            ws.write(row, 6, f"{r.get('ctr', 0):.2f}%", body_fmt)
            ws.write(row, 7, f"{r.get('vtr_2s', 0):.1f}%", body_fmt)
            ws.write(row, 8, f"€{r.get('spend', 0):,.0f}", body_fmt)
            row += 1

        # --- PLACEMENT SUMMARY ---
        row += 2
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  PLACEMENT SUMMARY",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#16a085",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        # Create placement summary pivot
        placement_col = (
            "placement_canonical" if "placement_canonical" in paid.columns else "format"
        )
        if placement_col in paid.columns:
            placement_summary = (
                paid.groupby(placement_col)
                .agg(
                    {
                        "composite_score": "mean",
                        "creative_name": "count",
                        "spend": "sum",
                        "vtr_2s": "mean",
                    }
                )
                .reset_index()
            )
            placement_summary = placement_summary.rename(
                columns={"creative_name": "n_creatives"}
            )
            placement_summary = placement_summary.sort_values(
                "composite_score", ascending=False
            )

            # Headers
            placement_headers = [
                "Placement",
                "Creatives",
                "Avg Score",
                "Total Spend",
                "Avg VTR",
            ]
            for ci, h in enumerate(placement_headers):
                ws.write(
                    row,
                    ci + 1,
                    h,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 9,
                            "font_color": "#888888",
                            "bottom": 1,
                            "bottom_color": "#dddddd",
                        }
                    ),
                )
            row += 1

            for _, r in placement_summary.head(8).iterrows():
                ws.write(row, 1, str(r[placement_col])[:18], body_fmt)
                ws.write(row, 2, int(r["n_creatives"]), body_fmt)
                score = r["composite_score"]
                if score >= 70:
                    ws.write(row, 3, f"{score:.1f}", tier_top_fmt)
                elif score >= 50:
                    ws.write(row, 3, f"{score:.1f}", tier_avg_fmt)
                else:
                    ws.write(row, 3, f"{score:.1f}", tier_below_fmt)
                ws.write(row, 4, f"€{r['spend']:,.0f}", body_fmt)
                ws.write(row, 5, f"{r['vtr_2s']:.1f}%", body_fmt)
                row += 1

        # --- OS PERFORMANCE (from raw ad-line data for clean splits) ---
        row += 2
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  OS PERFORMANCE",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#8e44ad",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        # Use raw data for OS breakdown (aggregated data has combined "Android, iOS" strings)
        raw_paid = (
            df_raw[df_raw["buying_type"] == "Paid"]
            if df_raw is not None and "buying_type" in df_raw.columns
            else pd.DataFrame()
        )
        if not raw_paid.empty and "os_target" in raw_paid.columns:
            os_summary = (
                raw_paid[raw_paid["os_target"].isin(["iOS", "Android", "Desktop"])]
                .groupby("os_target")
                .agg(
                    {
                        "composite_score": "mean",
                        "creative_name": "count",
                        "spend": "sum",
                        "engagement_rate": "mean",
                        "ctr": "mean",
                        "vtr_2s": "mean",
                    }
                )
                .reset_index()
                .rename(columns={"creative_name": "n_creatives"})
                .sort_values("composite_score", ascending=False)
            )
            os_headers = [
                "OS",
                "Ad Lines",
                "Avg Score",
                "Avg ER",
                "Avg CTR",
                "Avg VTR",
                "Total Spend",
            ]
            for ci, h in enumerate(os_headers):
                ws.write(
                    row,
                    ci + 1,
                    h,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 9,
                            "font_color": "#888888",
                            "bottom": 1,
                            "bottom_color": "#dddddd",
                        }
                    ),
                )
            row += 1
            for _, r in os_summary.iterrows():
                ws.write(row, 1, str(r["os_target"]), body_fmt)
                ws.write(row, 2, int(r["n_creatives"]), body_fmt)
                score = r["composite_score"]
                if score >= 70:
                    ws.write(row, 3, f"{score:.1f}", tier_top_fmt)
                elif score >= 50:
                    ws.write(row, 3, f"{score:.1f}", tier_avg_fmt)
                else:
                    ws.write(row, 3, f"{score:.1f}", tier_below_fmt)
                ws.write(row, 4, f"{r['engagement_rate']:.3f}%", body_fmt)
                ws.write(row, 5, f"{r['ctr']:.2f}%", body_fmt)
                ws.write(row, 6, f"{r['vtr_2s']:.1f}%", body_fmt)
                ws.write(row, 7, f"€{r['spend']:,.0f}", body_fmt)
                row += 1

        # --- OBJECTIVE × FORMAT MATRIX ---
        row += 1
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  OBJECTIVE × FORMAT MATRIX",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#16a085",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        # Create pivot of avg score by objective × format
        format_col = (
            "format_canonical" if "format_canonical" in paid.columns else "format"
        )
        obj_col = (
            "objective_normalized"
            if "objective_normalized" in paid.columns
            else "objective"
        )

        if format_col in paid.columns and obj_col in paid.columns:
            matrix = (
                paid.groupby([obj_col, format_col])
                .agg({"composite_score": "mean", "creative_name": "count"})
                .reset_index()
            )
            matrix = matrix.rename(columns={"creative_name": "n_creatives"})

            # Header row
            formats_in_data = sorted(matrix[format_col].dropna().unique())
            ws.write(row, 1, "Objective", bold_fmt)
            for fi, fmt_val in enumerate(formats_in_data):
                ws.write(row, 2 + fi, fmt_val, bold_fmt)
            row += 1

            # Data rows
            for obj_val in sorted(matrix[obj_col].dropna().unique()):
                ws.write(row, 1, str(obj_val)[:20], body_fmt)
                for fi, fmt_val in enumerate(formats_in_data):
                    cell_df = matrix[
                        (matrix[obj_col] == obj_val) & (matrix[format_col] == fmt_val)
                    ]
                    if not cell_df.empty:
                        score = cell_df["composite_score"].values[0]
                        n = int(cell_df["n_creatives"].values[0])
                        # Apply tier format based on score
                        if score >= 70:
                            ws.write(row, 2 + fi, f"{score:.0f} ({n})", tier_top_fmt)
                        elif score >= 50:
                            ws.write(row, 2 + fi, f"{score:.0f} ({n})", tier_avg_fmt)
                        else:
                            ws.write(row, 2 + fi, f"{score:.0f} ({n})", tier_below_fmt)
                    else:
                        ws.write(row, 2 + fi, "-", body_fmt)
                row += 1
        # --- ASSET TYPE COMPARISON (Brand vs Creator) ---
        row += 1
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  ASSET TYPE COMPARISON (Brand vs Creator)",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#e67e22",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        asset_col = (
            "asset_type_canonical" if "asset_type_canonical" in paid.columns else None
        )
        if asset_col and paid[asset_col].notna().any():
            asset_summary = (
                paid.groupby(asset_col)
                .agg(
                    {
                        "composite_score": "mean",
                        "creative_name": "count",
                        "spend": "sum",
                        "engagement_rate": "mean",
                        "ctr": "mean",
                        "vtr_2s": "mean",
                    }
                )
                .reset_index()
                .rename(columns={"creative_name": "n_creatives"})
                .sort_values("composite_score", ascending=False)
            )
            asset_headers = [
                "Asset Type",
                "Creatives",
                "Avg Score",
                "Avg ER",
                "Avg CTR",
                "Avg VTR",
                "Total Spend",
            ]
            for ci, h in enumerate(asset_headers):
                ws.write(
                    row,
                    ci + 1,
                    h,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 9,
                            "font_color": "#888888",
                            "bottom": 1,
                            "bottom_color": "#dddddd",
                        }
                    ),
                )
            row += 1
            for _, r in asset_summary.iterrows():
                ws.write(row, 1, str(r[asset_col]), body_fmt)
                ws.write(row, 2, int(r["n_creatives"]), body_fmt)
                score = r["composite_score"]
                if score >= 70:
                    ws.write(row, 3, f"{score:.1f}", tier_top_fmt)
                elif score >= 50:
                    ws.write(row, 3, f"{score:.1f}", tier_avg_fmt)
                else:
                    ws.write(row, 3, f"{score:.1f}", tier_below_fmt)
                ws.write(row, 4, f"{r['engagement_rate']:.3f}%", body_fmt)
                ws.write(row, 5, f"{r['ctr']:.2f}%", body_fmt)
                ws.write(row, 6, f"{r['vtr_2s']:.1f}%", body_fmt)
                ws.write(row, 7, f"€{r['spend']:,.0f}", body_fmt)
                row += 1

        # --- FORMAT COMPARISON (Video vs Static vs Motion) ---
        row += 2
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  FORMAT COMPARISON",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#2980b9",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        fmt_col = "format_canonical" if "format_canonical" in paid.columns else "format"
        if fmt_col in paid.columns and paid[fmt_col].notna().any():
            fmt_summary = (
                paid.groupby(fmt_col)
                .agg(
                    {
                        "composite_score": "mean",
                        "creative_name": "count",
                        "spend": "sum",
                        "engagement_rate": "mean",
                        "ctr": "mean",
                        "vtr_2s": "mean",
                    }
                )
                .reset_index()
                .rename(columns={"creative_name": "n_creatives"})
                .sort_values("composite_score", ascending=False)
            )
            fmt_headers = [
                "Format",
                "Creatives",
                "Avg Score",
                "Avg ER",
                "Avg CTR",
                "Avg VTR",
                "Total Spend",
            ]
            for ci, h in enumerate(fmt_headers):
                ws.write(
                    row,
                    ci + 1,
                    h,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 9,
                            "font_color": "#888888",
                            "bottom": 1,
                            "bottom_color": "#dddddd",
                        }
                    ),
                )
            row += 1
            for _, r in fmt_summary.iterrows():
                ws.write(row, 1, str(r[fmt_col]), body_fmt)
                ws.write(row, 2, int(r["n_creatives"]), body_fmt)
                score = r["composite_score"]
                if score >= 70:
                    ws.write(row, 3, f"{score:.1f}", tier_top_fmt)
                elif score >= 50:
                    ws.write(row, 3, f"{score:.1f}", tier_avg_fmt)
                else:
                    ws.write(row, 3, f"{score:.1f}", tier_below_fmt)
                ws.write(row, 4, f"{r['engagement_rate']:.3f}%", body_fmt)
                ws.write(row, 5, f"{r['ctr']:.2f}%", body_fmt)
                ws.write(row, 6, f"{r['vtr_2s']:.1f}%", body_fmt)
                ws.write(row, 7, f"€{r['spend']:,.0f}", body_fmt)
                row += 1

        # --- AUDIENCE SEGMENT COMPARISON ---
        row += 2
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  AUDIENCE SEGMENT COMPARISON",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#8e44ad",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        # Use raw data for audience segments, exploding comma-separated values
        seg_source = (
            raw_paid
            if not raw_paid.empty and "audience_segment" in raw_paid.columns
            else paid
        )
        if (
            "audience_segment" in seg_source.columns
            and seg_source["audience_segment"].notna().any()
        ):
            # Explode comma-separated segments ("R&F, RMKT") into individual rows
            seg_exploded = seg_source.copy()
            seg_exploded["audience_segment"] = (
                seg_exploded["audience_segment"].astype(str).str.split(r",\s*")
            )
            seg_exploded = seg_exploded.explode("audience_segment")
            seg_exploded["audience_segment"] = seg_exploded[
                "audience_segment"
            ].str.strip()
            seg_summary = (
                seg_exploded.groupby("audience_segment")
                .agg(
                    {
                        "composite_score": "mean",
                        "creative_name": "count",
                        "spend": "sum",
                        "engagement_rate": "mean",
                        "ctr": "mean",
                        "vtr_2s": "mean",
                    }
                )
                .reset_index()
                .rename(columns={"creative_name": "n_creatives"})
                .sort_values("composite_score", ascending=False)
            )
            # Filter out generic/empty segments
            seg_summary = seg_summary[
                ~seg_summary["audience_segment"].isin(["All", "", "nan"])
            ]

            if not seg_summary.empty:
                seg_headers = [
                    "Audience",
                    "Creatives",
                    "Avg Score",
                    "Avg ER",
                    "Avg CTR",
                    "Avg VTR",
                    "Total Spend",
                ]
                for ci, h in enumerate(seg_headers):
                    ws.write(
                        row,
                        ci + 1,
                        h,
                        workbook.add_format(
                            {
                                "bold": True,
                                "font_size": 9,
                                "font_color": "#888888",
                                "bottom": 1,
                                "bottom_color": "#dddddd",
                            }
                        ),
                    )
                row += 1
                for _, r in seg_summary.head(10).iterrows():
                    ws.write(row, 1, str(r["audience_segment"])[:25], body_fmt)
                    ws.write(row, 2, int(r["n_creatives"]), body_fmt)
                    score = r["composite_score"]
                    if score >= 70:
                        ws.write(row, 3, f"{score:.1f}", tier_top_fmt)
                    elif score >= 50:
                        ws.write(row, 3, f"{score:.1f}", tier_avg_fmt)
                    else:
                        ws.write(row, 3, f"{score:.1f}", tier_below_fmt)
                    ws.write(row, 4, f"{r['engagement_rate']:.3f}%", body_fmt)
                    ws.write(row, 5, f"{r['ctr']:.2f}%", body_fmt)
                    ws.write(row, 6, f"{r['vtr_2s']:.1f}%", body_fmt)
                    ws.write(row, 7, f"€{r['spend']:,.0f}", body_fmt)
                    row += 1

        # --- CAMPAIGN COMPARISON ---
        row += 2
        ws.merge_range(
            row,
            1,
            row,
            8,
            "  CAMPAIGN COMPARISON",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 14,
                    "font_color": "white",
                    "bg_color": "#34495e",
                    "valign": "vcenter",
                    "indent": 1,
                }
            ),
        )
        row += 1

        camp_col = (
            "campaign_normalized" if "campaign_normalized" in paid.columns else None
        )
        if camp_col and paid[camp_col].notna().any():
            camp_summary = (
                paid.groupby(camp_col)
                .agg(
                    {
                        "composite_score": "mean",
                        "creative_name": "count",
                        "spend": "sum",
                        "engagement_rate": "mean",
                        "ctr": "mean",
                        "vtr_2s": "mean",
                    }
                )
                .reset_index()
                .rename(columns={"creative_name": "n_creatives"})
                .sort_values("composite_score", ascending=False)
            )
            camp_summary = camp_summary[
                ~camp_summary[camp_col].isin(["Unknown", "", "nan"])
            ]

            if not camp_summary.empty:
                camp_headers = [
                    "Campaign",
                    "Creatives",
                    "Avg Score",
                    "Avg ER",
                    "Avg CTR",
                    "Avg VTR",
                    "Total Spend",
                ]
                for ci, h in enumerate(camp_headers):
                    ws.write(
                        row,
                        ci + 1,
                        h,
                        workbook.add_format(
                            {
                                "bold": True,
                                "font_size": 9,
                                "font_color": "#888888",
                                "bottom": 1,
                                "bottom_color": "#dddddd",
                            }
                        ),
                    )
                row += 1
                for _, r in camp_summary.head(15).iterrows():
                    ws.write(row, 1, str(r[camp_col])[:30], body_fmt)
                    ws.write(row, 2, int(r["n_creatives"]), body_fmt)
                    score = r["composite_score"]
                    if score >= 70:
                        ws.write(row, 3, f"{score:.1f}", tier_top_fmt)
                    elif score >= 50:
                        ws.write(row, 3, f"{score:.1f}", tier_avg_fmt)
                    else:
                        ws.write(row, 3, f"{score:.1f}", tier_below_fmt)
                    ws.write(row, 4, f"{r['engagement_rate']:.3f}%", body_fmt)
                    ws.write(row, 5, f"{r['ctr']:.2f}%", body_fmt)
                    ws.write(row, 6, f"{r['vtr_2s']:.1f}%", body_fmt)
                    ws.write(row, 7, f"€{r['spend']:,.0f}", body_fmt)
                    row += 1

        row += 2

        # --- KEY INSIGHTS ---
        ws.merge_range(row, 1, row, 8, "KEY INSIGHTS", section_fmt)
        row += 1

        # iOS vs Android winner
        if "os_target" in paid.columns:
            ios_df = paid[paid["os_target"] == "iOS"]
            android_df = paid[paid["os_target"] == "Android"]
            if not ios_df.empty and not android_df.empty:
                ios_score = ios_df["composite_score"].mean()
                android_score = android_df["composite_score"].mean()
                winner = "iOS" if ios_score > android_score else "Android"
                diff = abs(ios_score - android_score)
                ws.write(
                    row,
                    1,
                    f"• {winner} outperforms by {diff:.1f} pts avg score ({max(ios_score, android_score):.1f} vs {min(ios_score, android_score):.1f})",
                    body_fmt,
                )
                row += 1

        # Brand vs Creator winner
        asset_col = (
            "asset_type_canonical"
            if "asset_type_canonical" in paid.columns
            else "asset_type"
        )
        if asset_col in paid.columns:
            brand_df = paid[paid[asset_col] == "Brand"]
            creator_df = paid[paid[asset_col] == "Creator"]
            if not brand_df.empty and not creator_df.empty:
                brand_score = brand_df["composite_score"].mean()
                creator_score = creator_df["composite_score"].mean()
                winner = "Brand" if brand_score > creator_score else "Creator"
                diff = abs(brand_score - creator_score)
                ws.write(
                    row,
                    1,
                    f"• {winner} creatives outperform by {diff:.1f} pts avg score",
                    body_fmt,
                )
                row += 1

        # Platform performance
        if not paid.empty:
            tiktok_avg = (
                paid[paid["platform"] == "TikTok"]["composite_score"].mean()
                if "TikTok" in paid["platform"].values
                else 0
            )
            meta_avg = (
                paid[paid["platform"] == "Meta"]["composite_score"].mean()
                if "Meta" in paid["platform"].values
                else 0
            )
            if tiktok_avg > 0 and meta_avg > 0:
                platform_winner = "TikTok" if tiktok_avg > meta_avg else "Meta"
                ws.write(
                    row,
                    1,
                    f"• {platform_winner} platform shows higher avg scores ({max(tiktok_avg, meta_avg):.1f} vs {min(tiktok_avg, meta_avg):.1f})",
                    body_fmt,
                )
                row += 1

        row += 2

        # --- METHODOLOGY NOTE ---
        note_fmt = workbook.add_format(
            {
                "font_size": 10,
                "italic": True,
                "font_color": "#888888",
            }
        )
        ws.merge_range(
            row,
            1,
            row,
            8,
            "Methodology: Scores are percentile ranks within each Objective × Platform × Buying Type cohort. "
            "Paid and Boosting are scored separately — do not compare scores across cohorts. "
            "Native Attention Proxy replaces legacy brand measurement.",
            note_fmt,
        )

        # =====================================================================
        # HELPER SHEET: _DashboardData (Hidden, for formula references)
        # =====================================================================
        data_ws = workbook.add_worksheet("_DashboardData")
        data_ws.hide_gridlines(2)

        # Prepare columns for helper sheet
        helper_cols = [
            "creative_name",
            "platform",
            "format_canonical",
            "placement_canonical",
            "objective_normalized",
            "asset_type_canonical",
            "os_target",
            "campaign_normalized",
            "buying_type",
            "composite_score",
            "tier",
            "low_confidence",
            "vtr_2s",
            "completion_rate",
            "ctr",
            "engagement_rate",
            "spend",
            "reach",
            "frequency",
            "device_type",
        ]
        helper_cols = [c for c in helper_cols if c in all_data.columns]
        helper_df = all_data[helper_cols].copy()

        # Pre-sort by score descending so FILTER results are already ranked
        if "composite_score" in helper_df.columns:
            helper_df = helper_df.sort_values(
                "composite_score", ascending=False
            ).reset_index(drop=True)

        # Write header row
        for col_idx, col_name in enumerate(helper_cols):
            data_ws.write(0, col_idx, col_name, header_fmt)

        # Write data rows
        for row_idx, (_, row_data) in enumerate(helper_df.iterrows(), start=1):
            for col_idx, col_name in enumerate(helper_cols):
                val = row_data.get(col_name, "")
                if pd.isna(val):
                    val = ""
                elif isinstance(val, float):
                    val = round(val, 4)
                data_ws.write(row_idx, col_idx, val)

        # Set column widths
        data_ws.set_column("A:A", 45)  # creative_name
        data_ws.set_column("B:Z", 15)

        # Define named ranges for easier formula reference
        n_data_rows = len(helper_df)
        workbook.define_name(
            "_DataCreative", f"='_DashboardData'!$A$2:$A${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataPlatform", f"='_DashboardData'!$B$2:$B${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataFormat", f"='_DashboardData'!$C$2:$C${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataPlacement", f"='_DashboardData'!$D$2:$D${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataObjective", f"='_DashboardData'!$E$2:$E${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataAssetType", f"='_DashboardData'!$F$2:$F${n_data_rows + 1}"
        )
        workbook.define_name("_DataOS", f"='_DashboardData'!$G$2:$G${n_data_rows + 1}")
        workbook.define_name(
            "_DataCampaign", f"='_DashboardData'!$H$2:$H${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataBuyingType", f"='_DashboardData'!$I$2:$I${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataScore", f"='_DashboardData'!$J$2:$J${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataTier", f"='_DashboardData'!$K$2:$K${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataLowConf", f"='_DashboardData'!$L$2:$L${n_data_rows + 1}"
        )
        workbook.define_name("_DataVTR", f"='_DashboardData'!$M$2:$M${n_data_rows + 1}")
        workbook.define_name(
            "_DataCompletion", f"='_DashboardData'!$N$2:$N${n_data_rows + 1}"
        )
        workbook.define_name("_DataCTR", f"='_DashboardData'!$O$2:$O${n_data_rows + 1}")
        workbook.define_name(
            "_DataEngagement", f"='_DashboardData'!$P$2:$P${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataSpend", f"='_DashboardData'!$Q$2:$Q${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataReach", f"='_DashboardData'!$R$2:$R${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataFrequency", f"='_DashboardData'!$S$2:$S${n_data_rows + 1}"
        )
        workbook.define_name(
            "_DataDevice", f"='_DashboardData'!$T$2:$T${n_data_rows + 1}"
        )

        # Hide the helper sheet
        data_ws.hide()

        # =====================================================================
        # SHEET 2: HOW SCORING WORKS
        # =====================================================================
        ms = workbook.add_worksheet("How Scoring Works")
        ms.hide_gridlines(2)
        ms.set_tab_color("#2c3e50")
        ms.set_column("A:A", 3)
        ms.set_column("B:B", 22)
        ms.set_column("C:C", 35)
        ms.set_column("D:D", 25)
        ms.set_column("E:E", 25)
        ms.set_column("F:F", 25)
        ms.set_column("G:G", 25)
        ms.set_column("H:H", 25)

        mr = 1
        ms.merge_range(mr, 1, mr, 7, "How We Score Creatives", title_fmt)
        mr += 1
        ms.merge_range(
            mr,
            1,
            mr,
            7,
            "A plain-English guide to understanding the scores, adjustments, and attention proxy methodology",
            subtitle_fmt,
        )
        mr += 2

        # The Big Picture
        ms.merge_range(mr, 1, mr, 7, "The Big Picture", section_fmt)
        mr += 1
        ms.set_row(mr, 90)
        ms.merge_range(
            mr,
            1,
            mr,
            7,
            "Every creative gets a score out of 100. The score answers one question: "
            '"How well did this ad perform for what it was bought to do?"\n\n'
            "An ad bought for Video Views is judged mainly on whether people watched it. "
            "An ad bought for Engagement is judged on interaction rates. "
            "The score factors in budget significance, audience fatigue, and cost efficiency.\n\n"
            "IMPORTANT: Paid and Boosting creatives are scored in entirely separate cohorts. "
            "Do NOT compare a Paid score directly to a Boosting score — they rank against different peer groups.",
            body_fmt,
        )
        mr += 2

        # How to Read the Dashboard
        ms.merge_range(mr, 1, mr, 7, "How to Read the Dashboard", section_fmt)
        mr += 1
        ms.set_row(mr, 110)
        ms.merge_range(
            mr,
            1,
            mr,
            7,
            "The main dashboard is designed to answer one simple question: "
            '"What is performing best within the filters I have selected?"\n\n'
            "Each row in the Top Performers and Bottom Performers tables represents one creative within a specific "
            "Campaign + Platform + Buying Type + Format + Objective combination.\n\n"
            "This means the same creative name can appear more than once if it was used for different objectives. "
            "That is expected and helps avoid mixing performance from different buying goals.\n\n"
            "Placement labels are platform-aware in the dashboard, for example Meta Feed vs TikTok In Feed.\n\n"
            "KPI values on the dashboard are weighted rollups across all matching raw ad lines in that bucket. "
            "The second line under each creative name shows the Objective, Placement, Audience, and Source rows behind that "
            "rollup, so viewers can immediately see what is being combined.\n\n"
            "Use the Dimension section at the bottom of the dashboard to see raw-row level tactical performance when you split "
            "the filtered set by Asset Type, OS, Placement, Objective, or Audience Segment.",
            body_fmt,
        )
        mr += 2

        # What Matters Most
        ms.merge_range(mr, 1, mr, 7, "What Matters Most", section_fmt)
        mr += 1

        explainer_items = [
            (
                "Did people watch it?  (50%)",
                "View-through rate (VTR) is the #1 signal for VIDEO creatives. 2s VTR (TikTok) or 3s VTR (Meta) tells us "
                "whether the creative grabbed attention in the first moments. "
                "For STATIC assets, Engagement Rate and CTR replace VTR as the primary KPIs since view-through is not applicable.",
            ),
            (
                "Did they keep watching?  (25%)",
                "Completion rate, CTR, engagement, and shares are supporting signals for video. "
                "For static assets, share rate is the key secondary signal. "
                "They tell us whether people who were hooked then took action.",
            ),
            (
                "Was it cost-effective?  (15%)",
                "We compare each ad's cost per outcome against others with the same objective. "
                "Cheaper results = higher score.",
            ),
            (
                "Native Attention Quality  (10%)",
                "Native Attention Proxy uses platform-native metrics (hook rate, hold rate, completion patterns) to proxy "
                "attention quality. Only applies to video/motion assets. "
                "When inputs are unavailable, the score is renormalized across "
                "remaining components (55.6% / 27.8% / 16.7%).",
            ),
        ]
        for question, answer in explainer_items:
            ms.set_row(mr, 42)
            ms.write(mr, 1, question, bold_fmt)
            ms.merge_range(mr, 2, mr, 7, answer, body_fmt)
            mr += 1
        mr += 1

        # What We Adjust For
        ms.merge_range(mr, 1, mr, 7, "What We Adjust For", section_fmt)
        mr += 1
        adjustments = [
            (
                "Paid vs Boosting",
                "Paid ads (purpose-built, directly bought) and Boosting (organic content amplified with budget) "
                "are scored in entirely separate cohorts. A Boosting creative is always ranked against other "
                "Boosting creatives; a Paid creative against other Paid creatives. Do not compare scores across cohorts.",
            ),
            (
                "Ad fatigue",
                "If someone has seen the same ad more than twice, they're less likely to engage. "
                "Ads with high frequency get their score reduced.",
            ),
            (
                "Budget confidence",
                'Ads with less than €500 spend or under 10,000 reach are flagged as "low confidence" — '
                "results may not be reliable.",
            ),
            (
                "Video duration",
                "Completion rates are adjusted for duration. A 15s video with 3% completion isn't "
                "necessarily more engaging than a 90s video with 0.5% — short videos are just easier to finish.",
            ),
            (
                "Audience consistency",
                "A creative that works well across multiple targeting groups is genuinely strong. "
                "Inconsistent performers are moderated toward the group average.",
            ),
            (
                "Static vs Video scoring",
                "Static assets (images, carousels) are scored on Engagement Rate and CTR instead of VTR and completion rate, "
                "since view-through metrics do not apply. Video and static creatives are scored in separate cohorts "
                "so they are only ranked against peers of the same format.",
            ),
        ]
        for label, desc in adjustments:
            ms.set_row(mr, 45)
            ms.write(mr, 1, label, bold_fmt)
            ms.merge_range(mr, 2, mr, 7, desc, body_fmt)
            mr += 1
        mr += 1

        # Score Guide
        ms.merge_range(mr, 1, mr, 7, "Score Guide", section_fmt)
        mr += 1
        ms.write(
            mr,
            1,
            "Score",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 11,
                    "bg_color": "#1a1a2e",
                    "font_color": "white",
                }
            ),
        )
        ms.merge_range(
            mr,
            2,
            mr,
            4,
            "What It Means",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 11,
                    "bg_color": "#1a1a2e",
                    "font_color": "white",
                }
            ),
        )
        ms.merge_range(
            mr,
            5,
            mr,
            7,
            "What To Do",
            workbook.add_format(
                {
                    "bold": True,
                    "font_size": 11,
                    "bg_color": "#1a1a2e",
                    "font_color": "white",
                }
            ),
        )
        tiers = [
            ("85-100", "Top Performer", "Scale spend, use as template", tier_top_fmt),
            (
                "70-84",
                "Strong",
                "Keep running, consider budget increase",
                tier_strong_fmt,
            ),
            ("50-69", "Average", "Review for optimization", tier_avg_fmt),
            (
                "25-49",
                "Below Average",
                "Consider pausing, swap in stronger variant",
                tier_below_fmt,
            ),
            ("0-24", "Poor", "Pause and replace", tier_below_fmt),
        ]
        for score_range, tier_label, rec, fmt in tiers:
            mr += 1
            ms.write(mr, 1, score_range, fmt)
            ms.merge_range(mr, 2, mr, 4, tier_label, fmt)
            ms.merge_range(mr, 5, mr, 7, rec, fmt)

        # =====================================================================
        # SHEET 3: SUMMARY - TOP PERFORMERS
        # =====================================================================
        summary_df = (
            paid[~paid["low_confidence"]]
            .head(50)[[c for c in DASHBOARD_COLS if c in paid.columns]]
            .copy()
        )
        summary_ws = workbook.add_worksheet("Summary - Top Performers")
        summary_ws.hide_gridlines(2)
        summary_ws.set_tab_color("#16a085")

        summary_note = (
            f"Top 50 high-confidence Paid creatives (spend ≥ €500 & reach ≥ 10k). "
            f"Boosting creatives scored in separate cohort — see Boosting Rankings tab. "
            f"{total - high_conf} low-confidence Paid creatives excluded."
        )
        summary_ws.set_row(0, 22)
        summary_ws.merge_range(
            0,
            0,
            0,
            len(summary_df.columns) - 1,
            summary_note,
            workbook.add_format(
                {
                    "font_size": 10,
                    "italic": True,
                    "font_color": "#666666",
                    "text_wrap": True,
                }
            ),
        )
        summary_df.to_excel(
            writer, sheet_name="Summary - Top Performers", index=False, startrow=1
        )
        for col_num, value in enumerate(summary_df.columns):
            summary_ws.write(1, col_num, value, header_fmt)
        summary_ws.set_column("A:A", 42)
        summary_ws.set_column("B:Z", 14)
        summary_ws.autofilter(1, 0, len(summary_df) + 1, len(summary_df.columns) - 1)

        # Conditional formatting on score
        score_col_idx = (
            list(summary_df.columns).index("composite_score")
            if "composite_score" in summary_df.columns
            else None
        )
        if score_col_idx is not None:
            score_col_letter = chr(ord("A") + score_col_idx)
            summary_ws.conditional_format(
                f"{score_col_letter}3:{score_col_letter}{len(summary_df) + 2}",
                {
                    "type": "3_color_scale",
                    "min_color": "#f8d7da",
                    "mid_color": "#fff3cd",
                    "max_color": "#d4edda",
                    "min_type": "num",
                    "min_value": 0,
                    "mid_type": "num",
                    "mid_value": 50,
                    "max_type": "num",
                    "max_value": 100,
                },
            )

        # =====================================================================
        # SHEETS 4-7: TOP BY DIMENSION TABS
        # =====================================================================

        _write_top_by_dimension_sheet(
            writer,
            workbook,
            paid,
            "OS",
            "os_target",
            header_fmt,
            campaigns,
            target=target,
        )
        _write_top_by_dimension_sheet(
            writer,
            workbook,
            paid,
            "Asset Type",
            "asset_type_canonical"
            if "asset_type_canonical" in paid.columns
            else "asset_type",
            header_fmt,
            campaigns,
            target=target,
        )
        _write_top_by_dimension_sheet(
            writer,
            workbook,
            paid,
            "Placement",
            "placement_canonical"
            if "placement_canonical" in paid.columns
            else "format",
            header_fmt,
            campaigns,
            target=target,
        )
        _write_top_by_dimension_sheet(
            writer,
            workbook,
            paid,
            "Objective",
            "objective_normalized"
            if "objective_normalized" in paid.columns
            else "objective",
            header_fmt,
            campaigns,
            target=target,
        )
        _write_top_by_dimension_sheet(
            writer,
            workbook,
            paid,
            "Device",
            "device_type",
            header_fmt,
            campaigns,
            target=target,
        )

        # =====================================================================
        # SHEETS 8-11: PLATFORM RANKINGS (Paid)
        # =====================================================================

        for platform in ["TikTok", "Meta"]:
            platform_df = paid[paid["platform"] == platform]
            if platform_df.empty:
                continue
            _write_rankings_sheet(
                writer,
                workbook,
                platform_df,
                f"{platform} Paid Rankings",
                header_fmt,
                wrap_fmt,
            )

        # =====================================================================
        # SHEETS 12-13: PLATFORM RANKINGS (Boosting)
        # =====================================================================

        for platform in ["TikTok", "Meta"]:
            platform_df = (
                boosting[boosting["platform"] == platform]
                if not boosting.empty
                else pd.DataFrame()
            )
            if platform_df.empty:
                continue
            _write_rankings_sheet(
                writer,
                workbook,
                platform_df,
                f"{platform} Boosting Rankings",
                header_fmt,
                wrap_fmt,
            )

        # =====================================================================
        # SHEET 14: SPLITS ANALYSIS (Raw Variants)
        # =====================================================================

        if df_raw is not None and not df_raw.empty:
            _write_splits_sheet(writer, workbook, df_raw, header_fmt)
            _write_os_comparison_sheet(writer, workbook, df_raw, header_fmt)
            _write_asset_type_sheet(writer, workbook, df_raw, header_fmt)

        # =====================================================================
        # SHEET: CROSS-PLATFORM
        # =====================================================================

        cross = (
            paid[paid["cross_platform"] == True]
            if "cross_platform" in paid.columns
            else pd.DataFrame()
        )
        if not cross.empty:
            _write_cross_platform_sheet(writer, workbook, cross, header_fmt)

        # =====================================================================
        # SHEET: OBJ × FORMAT MATRIX
        # =====================================================================

        _write_objective_format_matrix(writer, workbook, paid, header_fmt)

        # =====================================================================
        # SHEET: LOOKER EXPORT
        # =====================================================================

        if df_raw is not None and not df_raw.empty:
            _write_looker_sheet(writer, workbook, df_raw, header_fmt)

        # =====================================================================
        # SHEET: RAW DATA (for debugging)
        # =====================================================================

        if df_raw is not None and not df_raw.empty:
            raw_cols = [
                c for c in df_raw.columns if c in LOOKER_COLS or c.endswith("_raw")
            ]
            raw_out = df_raw[raw_cols].copy() if raw_cols else df_raw.copy()
            raw_out.to_excel(writer, sheet_name="Raw Data", index=False)
            raw_ws = writer.sheets["Raw Data"]
            for i, col in enumerate(raw_out.columns):
                raw_ws.write(0, i, col, header_fmt)
            raw_ws.set_tab_color("#888888")

        # Set Dashboard as first active sheet
        ws.activate()

        writer.close()
        _rebuild_dashboard_with_openpyxl(
            temp_output_path,
            all_data,
            df_raw=df_raw,
            brand=brand,
            target=target,
        )
        os.replace(temp_output_path, output_path)
    finally:
        if os.path.exists(temp_output_path):
            os.remove(temp_output_path)

    print(f"Results exported to: {output_path}")


# =============================================================================
# OPENPYXL DASHBOARD REBUILD
# =============================================================================


def _clean_dashboard_label(val, default: str = "Unknown") -> str:
    """Normalize helper-sheet labels to plain strings."""
    if pd.isna(val):
        return default
    text = str(val).strip()
    if not text or text.lower() in {"nan", "none"}:
        return default
    if text.lower() == "na":
        return "Unknown"
    return text


def _build_display_name_lookup(df: pd.DataFrame) -> pd.DataFrame:
    """Append distinguishing context for creative names that appear multiple times."""
    key_cols = [
        "creative_name",
        "platform",
        "buying_type",
        "campaign_normalized",
        "objective_normalized",
        "format_canonical",
        "placement_canonical",
    ]
    display_df = df[key_cols].drop_duplicates().copy()
    duplicate_stats = display_df.groupby("creative_name", dropna=False).agg(
        n_rows=("creative_name", "size"),
        n_platforms=("platform", "nunique"),
        n_buying_types=("buying_type", "nunique"),
        n_campaigns=("campaign_normalized", "nunique"),
        n_objectives=("objective_normalized", "nunique"),
        n_formats=("format_canonical", "nunique"),
        n_placements=("placement_canonical", "nunique"),
    )
    duplicate_stats = duplicate_stats[duplicate_stats["n_rows"] > 1]
    duplicate_names = set(duplicate_stats.index.tolist())

    def _display_name(row: pd.Series) -> str:
        creative_name = _clean_dashboard_label(row["creative_name"], default="Unknown")
        if creative_name not in duplicate_names:
            return creative_name

        stats = duplicate_stats.loc[creative_name]
        context: list[str] = []
        context_rules = [
            ("objective_normalized", stats["n_objectives"] > 1),
            ("format_canonical", stats["n_formats"] > 1),
            ("placement_canonical", stats["n_placements"] > 1),
            ("platform", stats["n_platforms"] > 1),
            ("buying_type", stats["n_buying_types"] > 1),
            ("campaign_normalized", stats["n_campaigns"] > 1),
        ]

        for col_name, should_include in context_rules:
            if not should_include:
                continue
            value = _clean_dashboard_label(row[col_name], default="")
            if not value or value in {"All", "Unknown"}:
                continue
            if value.lower() in creative_name.lower():
                continue
            if value not in context:
                context.append(value)

        if not context:
            fallback = _clean_dashboard_label(row["objective_normalized"], default="")
            if fallback:
                context.append(fallback)

        return f"{creative_name} [{' | '.join(context[:4])}]"

    display_df["display_name"] = display_df.apply(_display_name, axis=1)
    return display_df


def _with_campaign_scope_rows(base_df: pd.DataFrame) -> pd.DataFrame:
    """Return exact-campaign rows plus an All-campaign copy."""
    exact_rows = base_df.rename(
        columns={"campaign_normalized": "campaign_scope"}
    ).copy()
    all_rows = base_df.copy()
    all_rows.insert(0, "campaign_scope", "All")
    return pd.concat([all_rows, exact_rows], ignore_index=True)


def _dashboard_number_format(metric_type: str) -> str:
    """Central number-format mapping used across dashboard and support sheets."""
    return {
        "score": "0.0",
        "currency": "€#,##0.00",
        "count": "#,##0",
        "rate": '0.00"%"',
        "frequency": '0.00"x"',
    }.get(metric_type, "General")


def _append_dashboard_context(label: str, context_items: list[str]) -> str:
    """Append dashboard context tags while avoiding duplicate bracket fragments."""
    clean_label = _clean_dashboard_label(label, default="Unknown")
    additions = []
    existing_lower: set[str] = set()

    match = re.match(r"^(.*)\s\[(.*)\]$", clean_label)
    if match:
        base_label = match.group(1).strip()
        existing_items = [
            item.strip() for item in match.group(2).split("|") if item.strip()
        ]
        existing_lower = {item.lower() for item in existing_items}
    else:
        base_label = clean_label
        existing_items = []

    for item in context_items:
        clean_item = _clean_dashboard_label(item, default="")
        if not clean_item:
            continue
        if clean_item.lower() in existing_lower:
            continue
        additions.append(clean_item)
        existing_lower.add(clean_item.lower())

    merged_items = existing_items + additions
    if not merged_items:
        return base_label
    return f"{base_label} [{' | '.join(merged_items)}]"


def _platform_placement_label(platform: str, placement: str) -> str:
    """Make placement labels explicitly platform-aware for non-technical viewers."""
    clean_platform = _clean_dashboard_label(platform, default="")
    clean_placement = _clean_dashboard_label(placement, default="Unknown")
    if not clean_platform or clean_platform in {"All", "Unknown"}:
        return clean_placement
    if clean_placement.lower().startswith(clean_platform.lower()):
        return clean_placement
    return f"{clean_platform} {clean_placement}"


def _extract_delivery_line_label(ad_name_raw) -> str:
    """Extract a short delivery-line identifier for raw-row level drilldowns."""
    text = _clean_dashboard_label(ad_name_raw, default="")
    if not text:
        return ""

    cell_match = re.search(r"(Cell\d+)", text, flags=re.IGNORECASE)
    if cell_match:
        return cell_match.group(1)

    opid_match = re.search(r"(OPID-\d+)", text, flags=re.IGNORECASE)
    if opid_match:
        return opid_match.group(1)

    return ""


def _detail_audience_label(audience_segment: str, ad_name_raw) -> str:
    """Build a readable audience/tactic label for raw-row detail views."""
    audience = _clean_dashboard_label(audience_segment, default="Unknown")
    line_label = _extract_delivery_line_label(ad_name_raw)
    if not line_label:
        return audience
    if audience in {"", "Unknown"}:
        return line_label
    return f"{audience} / {line_label}"


def _build_dimension_display_name(
    base_name: str, dimension_name: str, dimension_value: str
) -> str:
    """Tag grouped-output creative labels with the active dimension value."""
    label = _clean_dashboard_label(base_name, default="Unknown")
    dim_value = _clean_dashboard_label(dimension_value, default="")
    if not dim_value:
        return label
    return _append_dashboard_context(label, [dim_value])


def _header_metric_type(header) -> str | None:
    """Infer workbook number format from a header label."""
    if header is None:
        return None

    text = str(header).strip().lower()
    normalized = (
        text.replace("×", "x").replace(" ", "_").replace("/", "_").replace("-", "_")
    )

    if normalized in {
        "score",
        "avg_score",
        "composite_score",
        "attention_proxy_score",
    } or normalized.startswith("composite_score_"):
        return "score"
    if normalized in {
        "spend",
        "total_spend",
        "cost_per_complete_view",
        "cost_per_click",
        "cpm",
    } or normalized.startswith("spend_"):
        return "currency"
    if normalized in {
        "reach",
        "impressions",
        "clicks",
        "video_views_100",
        "shares",
        "engagements",
        "n_variants",
        "n_campaigns",
        "group_size",
        "rank_in_group",
        "creatives",
        "creative_count",
        "n_creatives",
        "count",
        "filtered_count",
        "total_high_confidence",
        "creatives_count",
        "rank_desc",
        "source_rows",
    } or normalized.startswith(
        ("reach_", "impressions_", "clicks_", "shares_", "engagements_")
    ):
        return "count"
    if normalized in {"frequency"} or normalized.startswith("frequency_"):
        return "frequency"
    if normalized in {
        "vtr",
        "avg_vtr",
        "vtr_2s",
        "hook_rate",
        "hold_rate",
        "completion_rate",
        "completion_vs_expected",
        "engagement_rate",
        "ctr",
        "share_rate",
        "canonical_hook_rate",
        "canonical_hold_rate",
        "canonical_completion_rate",
        "canonical_engagement_rate",
        "canonical_ctr",
    } or normalized.startswith(
        (
            "vtr_",
            "vtr_2s_",
            "completion_rate_",
            "engagement_rate_",
            "ctr_",
            "share_rate_",
        )
    ):
        return "rate"
    return None


def _apply_header_number_formats(ws, header_row: int = 1):
    """Apply metric-aware number formats to a worksheet based on its header row."""
    for col_idx in range(1, ws.max_column + 1):
        metric_type = _header_metric_type(ws.cell(header_row, col_idx).value)
        if not metric_type:
            continue
        number_format = _dashboard_number_format(metric_type)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value in (None, ""):
                continue
            if isinstance(cell.value, str) and not cell.data_type == "f":
                continue
            cell.number_format = number_format


def _aggregate_dashboard_group(group: pd.DataFrame) -> pd.Series:
    """Aggregate raw rows into a single dashboard-ready record."""
    impressions = float(group["impressions"].sum())
    reach = float(group["reach"].sum())
    spend = float(group["spend"].sum())
    clicks = float(group["clicks"].sum())
    engagements = float(group["engagements"].sum())
    video_views_100 = float(group["video_views_100"].sum())
    weighted_vtr = (
        (group["vtr_2s"].fillna(0) * group["impressions"].fillna(0)).sum() / impressions
        if impressions > 0
        else float(group["vtr_2s"].mean())
    )
    creative_name = _clean_dashboard_label(
        group["creative_name"].iloc[0], default="Unknown"
    )
    display_name = (
        _clean_dashboard_label(group["display_name"].iloc[0], default=creative_name)
        if "display_name" in group.columns
        else creative_name
    )
    placement_values = sorted(
        {
            _clean_dashboard_label(value, default="Unknown")
            for value in group.get(
                "placement_display",
                group.get("placement_canonical", pd.Series(dtype=object)),
            ).tolist()
            if _clean_dashboard_label(value, default="Unknown")
        }
    )
    if len(placement_values) == 0:
        placement_label = "Unknown"
    elif len(placement_values) == 1:
        placement_label = placement_values[0]
    elif len(placement_values) <= 3:
        placement_label = "Multiple: " + ", ".join(placement_values)
    else:
        placement_label = "Multiple"
    audience_values = sorted(
        {
            _clean_dashboard_label(value, default="Unknown")
            for value in group.get("audience_segment", pd.Series(dtype=object)).tolist()
            if _clean_dashboard_label(value, default="Unknown")
        }
    )
    if len(audience_values) == 0:
        audience_label = "Unknown"
    elif len(audience_values) == 1:
        audience_label = audience_values[0]
    elif len(audience_values) <= 3:
        audience_label = "Mixed: " + ", ".join(audience_values)
    else:
        audience_label = "Mixed"
    source_rows = int(len(group))

    return pd.Series(
        {
            "display_name": display_name,
            "placement_canonical": placement_label,
            "audience_segment_label": audience_label,
            "source_rows": source_rows,
            "composite_score": float(group["composite_score"].mean()),
            "vtr_2s": weighted_vtr if pd.notna(weighted_vtr) else 0.0,
            "completion_rate": (video_views_100 / impressions) * 100
            if impressions > 0
            else float(group["completion_rate"].mean()),
            "ctr": (clicks / impressions) * 100
            if impressions > 0
            else float(group["ctr"].mean()),
            "engagement_rate": (engagements / impressions) * 100
            if impressions > 0
            else float(group["engagement_rate"].mean()),
            "spend": spend,
            "reach": reach,
            "impressions": impressions,
            "frequency": (impressions / reach)
            if reach > 0
            else float(group["frequency"].mean()),
            "low_confidence": bool(group["low_confidence"].iloc[0]),
            "tier": _clean_dashboard_label(group["tier"].iloc[0], default="Average"),
        }
    )


def _build_campaign_scoped_rows(
    df: pd.DataFrame, group_cols: list[str]
) -> pd.DataFrame:
    """Build exact-campaign and all-campaign helper rows from the same source data."""
    exact = (
        df.groupby(["campaign_normalized"] + group_cols, dropna=False)
        .apply(_aggregate_dashboard_group)
        .reset_index()
        .rename(columns={"campaign_normalized": "campaign_scope"})
    )

    all_rows = (
        df.groupby(group_cols, dropna=False)
        .apply(_aggregate_dashboard_group)
        .reset_index()
    )
    all_rows.insert(0, "campaign_scope", "All")

    scoped = pd.concat([all_rows, exact], ignore_index=True)
    return scoped.fillna(
        {
            "vtr_2s": 0.0,
            "completion_rate": 0.0,
            "ctr": 0.0,
            "engagement_rate": 0.0,
            "spend": 0.0,
            "reach": 0.0,
            "impressions": 0.0,
            "frequency": 0.0,
            "composite_score": 0.0,
            "display_name": "",
            "tier": "Average",
            "low_confidence": False,
        }
    )


def _build_dashboard_support_data(
    df: pd.DataFrame, df_raw: "pd.DataFrame | None"
) -> dict:
    """Prepare normalized helper datasets for the selector-driven dashboard."""
    objective_col = (
        "objective_normalized" if "objective_normalized" in df.columns else "objective"
    )
    df = df.copy()
    df[objective_col] = df[objective_col].apply(_clean_dashboard_label)
    df["objective_normalized"] = df[objective_col]
    for col, default in [
        ("campaign_normalized", "Unknown"),
        ("platform", "Unknown"),
        ("buying_type", "Paid"),
        ("format_canonical", "Unknown"),
        ("placement_canonical", "Unknown"),
        ("creative_name", "Unknown"),
    ]:
        if col not in df.columns:
            df[col] = default
        df[col] = df[col].apply(lambda x: _clean_dashboard_label(x, default=default))
    display_lookup = _build_display_name_lookup(df)
    df = df.merge(
        display_lookup,
        on=[
            "creative_name",
            "platform",
            "buying_type",
            "campaign_normalized",
            "objective_normalized",
            "format_canonical",
            "placement_canonical",
        ],
        how="left",
    )
    df["display_name"] = df["display_name"].fillna(df["creative_name"])

    raw = df_raw.copy() if df_raw is not None else pd.DataFrame()
    if raw.empty:
        raw = df.copy()

    raw_objective_col = (
        "objective_normalized" if "objective_normalized" in raw.columns else "objective"
    )
    raw[raw_objective_col] = raw[raw_objective_col].apply(_clean_dashboard_label)
    raw["objective_normalized"] = raw[raw_objective_col]

    for col, default in [
        ("campaign_normalized", "Unknown"),
        ("platform", "Unknown"),
        ("buying_type", "Paid"),
        ("format_canonical", "Unknown"),
        ("placement_canonical", "Unknown"),
        ("asset_type_canonical", "Unknown"),
        ("os_target", "Unknown"),
        ("creative_name", "Unknown"),
    ]:
        if col not in raw.columns:
            raw[col] = default
        raw[col] = raw[col].apply(lambda x: _clean_dashboard_label(x, default=default))

    for numeric_col in [
        "impressions",
        "reach",
        "spend",
        "clicks",
        "engagements",
        "video_views_100",
        "vtr_2s",
        "completion_rate",
        "ctr",
        "engagement_rate",
        "frequency",
    ]:
        if numeric_col not in raw.columns:
            raw[numeric_col] = 0.0
        raw[numeric_col] = pd.to_numeric(raw[numeric_col], errors="coerce").fillna(0.0)

    if "low_confidence" not in raw.columns:
        raw["low_confidence"] = False
    raw["low_confidence"] = raw["low_confidence"].fillna(False).astype(bool)

    merge_cols = [
        "creative_name",
        "platform",
        "buying_type",
        "campaign_normalized",
        raw_objective_col,
    ]
    creative_lookup_cols = merge_cols + [
        "display_name",
        "composite_score",
        "tier",
        "low_confidence",
    ]
    creative_lookup = (
        df[creative_lookup_cols]
        .drop_duplicates()
        .rename(columns={objective_col: raw_objective_col})
    )

    raw_enriched = raw.merge(
        creative_lookup,
        on=merge_cols,
        how="left",
        suffixes=("", "_creative"),
    )
    raw_enriched["raw_variant_score"] = pd.to_numeric(
        raw_enriched.get("composite_score"), errors="coerce"
    ).fillna(0.0)
    if "display_name_creative" in raw_enriched.columns:
        raw_enriched["display_name"] = raw_enriched["display_name_creative"].fillna(
            raw_enriched.get("display_name")
        )
    raw_enriched["display_name"] = raw_enriched.get(
        "display_name", raw_enriched["creative_name"]
    ).fillna(raw_enriched["creative_name"])
    if "composite_score_creative" in raw_enriched.columns:
        raw_enriched["composite_score"] = raw_enriched[
            "composite_score_creative"
        ].fillna(raw_enriched.get("composite_score"))
    raw_enriched["composite_score"] = pd.to_numeric(
        raw_enriched.get("composite_score"), errors="coerce"
    ).fillna(0.0)
    if "tier_creative" in raw_enriched.columns:
        raw_enriched["tier"] = raw_enriched["tier_creative"].fillna(
            raw_enriched.get("tier")
        )
    raw_enriched["tier"] = raw_enriched.get("tier", "Average").fillna("Average")
    if "low_confidence_creative" in raw_enriched.columns:
        raw_enriched["low_confidence"] = (
            raw_enriched["low_confidence_creative"].fillna(False).astype(bool)
        )
    else:
        raw_enriched["low_confidence"] = (
            raw_enriched["low_confidence"].fillna(False).astype(bool)
        )
    raw_enriched["objective_normalized"] = raw_enriched[raw_objective_col].apply(
        _clean_dashboard_label
    )
    raw_enriched["placement_display"] = raw_enriched.apply(
        lambda row: _platform_placement_label(
            row.get("platform", ""), row.get("placement_canonical", "")
        ),
        axis=1,
    )
    raw_enriched["audience_detail_label"] = raw_enriched.apply(
        lambda row: _detail_audience_label(
            row.get("audience_segment", ""), row.get("ad_name_raw", "")
        ),
        axis=1,
    )

    def _expand_selector_scopes(
        base_df: pd.DataFrame, include_format_all: bool = True
    ) -> pd.DataFrame:
        frames = []
        format_options = [False, True] if include_format_all else [False]
        for include_all_platform in [False, True]:
            for include_both_buying in [False, True]:
                for include_all_format in format_options:
                    scoped = base_df.copy()
                    scoped["platform"] = (
                        "All" if include_all_platform else scoped["platform"]
                    )
                    scoped["buying_type"] = (
                        "Both" if include_both_buying else scoped["buying_type"]
                    )
                    if include_format_all:
                        scoped["format_canonical"] = (
                            "All" if include_all_format else scoped["format_canonical"]
                        )
                    frames.append(scoped)
        return pd.concat(frames, ignore_index=True)

    creative_group_cols = [
        "creative_name",
        "platform",
        "buying_type",
        "objective_normalized",
        "format_canonical",
    ]
    creative_rows_exact = _build_campaign_scoped_rows(raw_enriched, creative_group_cols)
    creative_rows_exact = creative_rows_exact[
        [
            "campaign_scope",
            "platform",
            "buying_type",
            "format_canonical",
            "creative_name",
            "objective_normalized",
            "placement_canonical",
            "audience_segment_label",
            "source_rows",
            "composite_score",
            "vtr_2s",
            "completion_rate",
            "ctr",
            "engagement_rate",
            "spend",
            "reach",
            "frequency",
            "display_name",
            "impressions",
            "low_confidence",
            "tier",
        ]
    ].sort_values(
        ["campaign_scope", "platform", "buying_type", "composite_score"],
        ascending=[True, True, True, False],
    )
    creative_rows = _expand_selector_scopes(creative_rows_exact).reset_index(drop=True)
    creative_rows["_row_id"] = np.arange(len(creative_rows))
    creative_scope_cols = [
        "campaign_scope",
        "platform",
        "buying_type",
        "format_canonical",
    ]
    creative_desc = creative_rows.sort_values(
        creative_scope_cols + ["composite_score", "creative_name", "_row_id"],
        ascending=[True, True, True, True, False, True, True],
    ).copy()
    creative_desc["rank_desc"] = (
        creative_desc.groupby(creative_scope_cols).cumcount() + 1
    )
    creative_asc = creative_rows.sort_values(
        creative_scope_cols + ["composite_score", "creative_name", "_row_id"],
        ascending=[True, True, True, True, True, True, True],
    ).copy()
    creative_asc["rank_asc"] = creative_asc.groupby(creative_scope_cols).cumcount() + 1
    creative_rows = creative_rows.merge(
        creative_desc[["_row_id", "rank_desc"]], on="_row_id", how="left"
    )
    creative_rows = creative_rows.merge(
        creative_asc[["_row_id", "rank_asc"]], on="_row_id", how="left"
    )
    creative_rows["key_desc"] = (
        creative_rows["campaign_scope"]
        + "|"
        + creative_rows["platform"]
        + "|"
        + creative_rows["buying_type"]
        + "|"
        + creative_rows["format_canonical"]
        + "|"
        + creative_rows["rank_desc"].astype(int).astype(str)
    )
    creative_rows["key_asc"] = (
        creative_rows["campaign_scope"]
        + "|"
        + creative_rows["platform"]
        + "|"
        + creative_rows["buying_type"]
        + "|"
        + creative_rows["format_canonical"]
        + "|"
        + creative_rows["rank_asc"].astype(int).astype(str)
    )
    creative_rows = creative_rows.sort_values(
        creative_scope_cols + ["rank_desc", "creative_name"],
        ascending=[True, True, True, True, True, True],
    ).drop(columns=["_row_id"])

    dimension_frames = []
    dimension_map = [
        ("Asset Type", "asset_type_canonical"),
        ("Placement", "placement_display"),
        ("OS", "os_target"),
        ("Device", "device_type"),
        ("Objective", "objective_normalized"),
        ("Audience Segment", "audience_segment"),
    ]
    dim_group_cols = [
        "creative_name",
        "platform",
        "buying_type",
        "objective_normalized",
        "format_canonical",
    ]

    for dimension_name, source_col in dimension_map:
        subset = raw_enriched.copy()
        subset["dimension_name"] = dimension_name
        # Audience segments can be comma-separated ("R&F, RMKT") — explode to individual rows
        if dimension_name == "Audience Segment":
            subset["dimension_value"] = (
                subset[source_col].astype(str).str.split(r",\s*")
            )
            subset = subset.explode("dimension_value")
            subset["dimension_value"] = subset["dimension_value"].apply(
                _clean_dashboard_label
            )
        else:
            subset["dimension_value"] = subset[source_col].apply(_clean_dashboard_label)
        if dimension_name == "OS":
            subset = subset[subset["dimension_value"].isin(["iOS", "Android"])]
        if dimension_name == "Device":
            subset = subset[subset["dimension_value"].isin(["Mobile", "Desktop"])]
        subset = subset[subset["dimension_value"] != "Unknown"]
        scoped = _build_campaign_scoped_rows(
            subset,
            ["dimension_name", "dimension_value"] + dim_group_cols,
        )
        dimension_frames.append(scoped)

    dimension_rows_exact = pd.concat(dimension_frames, ignore_index=True)
    dimension_rows_exact["display_name"] = dimension_rows_exact.apply(
        lambda row: _build_dimension_display_name(
            row.get("display_name", row.get("creative_name", "")),
            row.get("dimension_name", ""),
            row.get("dimension_value", ""),
        ),
        axis=1,
    )
    dimension_rows_exact = dimension_rows_exact[
        [
            "campaign_scope",
            "dimension_name",
            "dimension_value",
            "platform",
            "buying_type",
            "format_canonical",
            "creative_name",
            "objective_normalized",
            "placement_canonical",
            "audience_segment_label",
            "source_rows",
            "composite_score",
            "vtr_2s",
            "completion_rate",
            "ctr",
            "engagement_rate",
            "spend",
            "reach",
            "frequency",
            "display_name",
            "impressions",
            "low_confidence",
        ]
    ].sort_values(
        [
            "dimension_name",
            "campaign_scope",
            "platform",
            "buying_type",
            "dimension_value",
            "composite_score",
        ],
        ascending=[True, True, True, True, True, False],
    )
    dimension_rows = _expand_selector_scopes(dimension_rows_exact).reset_index(
        drop=True
    )
    dimension_rows["_row_id"] = np.arange(len(dimension_rows))
    dim_scope_cols = [
        "campaign_scope",
        "dimension_name",
        "dimension_value",
        "platform",
        "buying_type",
        "format_canonical",
    ]
    dim_desc = dimension_rows.sort_values(
        dim_scope_cols + ["composite_score", "creative_name", "_row_id"],
        ascending=[True, True, True, True, True, True, False, True, True],
    ).copy()
    dim_desc["rank_desc"] = dim_desc.groupby(dim_scope_cols).cumcount() + 1
    dimension_rows = dimension_rows.merge(
        dim_desc[["_row_id", "rank_desc"]], on="_row_id", how="left"
    )
    dimension_rows["group_key"] = (
        dimension_rows["campaign_scope"]
        + "|"
        + dimension_rows["dimension_name"]
        + "|"
        + dimension_rows["platform"]
        + "|"
        + dimension_rows["buying_type"]
        + "|"
        + dimension_rows["format_canonical"]
        + "|"
        + dimension_rows["dimension_value"]
        + "|"
        + dimension_rows["rank_desc"].astype(int).astype(str)
    )
    dimension_rows = dimension_rows.sort_values(
        [
            "dimension_name",
            "campaign_scope",
            "platform",
            "buying_type",
            "format_canonical",
            "dimension_value",
            "rank_desc",
        ],
        ascending=[True, True, True, True, True, True, True],
    ).drop(columns=["_row_id"])
    scoped_dimensions = dimension_rows[~dimension_rows["low_confidence"]].copy()

    detail_dimension_frames = []
    for dimension_name, source_col in dimension_map:
        subset = raw_enriched.copy()
        subset["dimension_name"] = dimension_name
        # Audience segments can be comma-separated ("R&F, RMKT") — explode to individual rows
        if dimension_name == "Audience Segment":
            subset["dimension_value"] = (
                subset[source_col].astype(str).str.split(r",\s*")
            )
            subset = subset.explode("dimension_value")
            subset["dimension_value"] = subset["dimension_value"].apply(
                _clean_dashboard_label
            )
        else:
            subset["dimension_value"] = subset[source_col].apply(_clean_dashboard_label)
        if dimension_name == "OS":
            subset = subset[subset["dimension_value"].isin(["iOS", "Android"])]
        if dimension_name == "Device":
            subset = subset[subset["dimension_value"].isin(["Mobile", "Desktop"])]
        subset = subset[subset["dimension_value"] != "Unknown"]
        detail_rows = (
            subset[
                [
                    "campaign_normalized",
                    "dimension_name",
                    "dimension_value",
                    "platform",
                    "buying_type",
                    "format_canonical",
                    "creative_name",
                    "objective_normalized",
                    "placement_display",
                    "audience_detail_label",
                    "raw_variant_score",
                    "vtr_2s",
                    "completion_rate",
                    "ctr",
                    "engagement_rate",
                    "spend",
                    "reach",
                    "frequency",
                    "display_name",
                    "impressions",
                    "low_confidence",
                ]
            ]
            .copy()
            .rename(
                columns={
                    "campaign_normalized": "campaign_scope",
                    "placement_display": "placement_canonical",
                    "audience_detail_label": "audience_segment_label",
                    "raw_variant_score": "composite_score",
                }
            )
        )
        detail_rows["source_rows"] = 1
        detail_rows = detail_rows[
            [
                "campaign_scope",
                "dimension_name",
                "dimension_value",
                "platform",
                "buying_type",
                "format_canonical",
                "creative_name",
                "objective_normalized",
                "placement_canonical",
                "audience_segment_label",
                "source_rows",
                "composite_score",
                "vtr_2s",
                "completion_rate",
                "ctr",
                "engagement_rate",
                "spend",
                "reach",
                "frequency",
                "display_name",
                "impressions",
                "low_confidence",
            ]
        ]
        detail_dimension_frames.append(
            pd.concat(
                [
                    detail_rows.assign(campaign_scope="All"),
                    detail_rows,
                ],
                ignore_index=True,
            )
        )

    dimension_detail_rows_exact = pd.concat(detail_dimension_frames, ignore_index=True)
    dimension_detail_rows = _expand_selector_scopes(
        dimension_detail_rows_exact
    ).reset_index(drop=True)
    dimension_detail_rows["_row_id"] = np.arange(len(dimension_detail_rows))
    detail_scope_cols = [
        "campaign_scope",
        "dimension_name",
        "dimension_value",
        "platform",
        "buying_type",
        "format_canonical",
    ]
    detail_desc = dimension_detail_rows.sort_values(
        detail_scope_cols
        + [
            "composite_score",
            "creative_name",
            "objective_normalized",
            "audience_segment_label",
            "_row_id",
        ],
        ascending=[True, True, True, True, True, True, False, True, True, True, True],
    ).copy()
    dimension_detail_rows["rank_desc"] = (
        detail_desc.groupby(detail_scope_cols).cumcount() + 1
    )
    dimension_detail_rows["group_key"] = (
        dimension_detail_rows["campaign_scope"]
        + "|"
        + dimension_detail_rows["dimension_name"]
        + "|"
        + dimension_detail_rows["platform"]
        + "|"
        + dimension_detail_rows["buying_type"]
        + "|"
        + dimension_detail_rows["format_canonical"]
        + "|"
        + dimension_detail_rows["dimension_value"]
        + "|"
        + dimension_detail_rows["rank_desc"].astype(int).astype(str)
    )
    dimension_detail_rows = dimension_detail_rows.sort_values(
        [
            "dimension_name",
            "campaign_scope",
            "platform",
            "buying_type",
            "format_canonical",
            "dimension_value",
            "rank_desc",
        ],
        ascending=[True, True, True, True, True, True, True],
    ).drop(columns=["_row_id"])

    def _build_dimension_summary(dimension_name: str) -> pd.DataFrame:
        dim_df = scoped_dimensions[
            scoped_dimensions["dimension_name"] == dimension_name
        ].copy()
        scope_cols = [
            "campaign_scope",
            "platform",
            "buying_type",
            "format_canonical",
            "dimension_value",
        ]
        creative_level = (
            dim_df.groupby(scope_cols + ["creative_name"], dropna=False)
            .agg(creative_score=("composite_score", "mean"))
            .reset_index()
        )
        score_summary = (
            creative_level.groupby(scope_cols, dropna=False)
            .agg(
                creatives_count=("creative_name", "nunique"),
                avg_score=("creative_score", "mean"),
            )
            .reset_index()
        )
        metric_summary = (
            dim_df.groupby(scope_cols, dropna=False)
            .agg(
                total_spend=("spend", "sum"),
                total_impressions=("impressions", "sum"),
            )
            .reset_index()
        )
        vtr_summary = (
            dim_df.groupby(scope_cols, dropna=False)
            .apply(
                lambda grp: (
                    (grp["vtr_2s"] * grp["impressions"]).sum()
                    / grp["impressions"].sum()
                    if grp["impressions"].sum()
                    else 0.0
                ),
                include_groups=False,
            )
            .reset_index(name="avg_vtr")
        )
        summary = (
            score_summary.merge(
                metric_summary.drop(columns=["total_impressions"]),
                on=scope_cols,
                how="left",
            )
            .merge(vtr_summary, on=scope_cols, how="left")
            .rename(columns={"dimension_value": "label"})
        )
        summary = summary.reset_index(drop=True)
        summary["_row_id"] = np.arange(len(summary))
        rank_scope_cols = [
            "campaign_scope",
            "platform",
            "buying_type",
            "format_canonical",
        ]
        ranked = summary.sort_values(
            rank_scope_cols + ["avg_score", "label", "_row_id"],
            ascending=[True, True, True, True, False, True, True],
        ).copy()
        ranked["rank_desc"] = ranked.groupby(rank_scope_cols).cumcount() + 1
        summary = summary.merge(
            ranked[["_row_id", "rank_desc"]], on="_row_id", how="left"
        )
        summary["summary_key"] = (
            summary["campaign_scope"]
            + "|"
            + summary["platform"]
            + "|"
            + summary["buying_type"]
            + "|"
            + summary["format_canonical"]
            + "|"
            + summary["rank_desc"].astype(int).astype(str)
        )
        return summary.sort_values(rank_scope_cols + ["rank_desc"]).drop(
            columns=["_row_id"]
        )

    placement_summary = _build_dimension_summary("Placement")
    os_summary = _build_dimension_summary("OS")
    asset_summary = _build_dimension_summary("Asset Type")

    dimension_label_rows = (
        scoped_dimensions[
            [
                "campaign_scope",
                "dimension_name",
                "dimension_value",
                "platform",
                "buying_type",
                "format_canonical",
            ]
        ]
        .drop_duplicates()
        .sort_values(
            [
                "dimension_name",
                "campaign_scope",
                "platform",
                "buying_type",
                "format_canonical",
                "dimension_value",
            ]
        )
        .reset_index(drop=True)
    )
    dimension_label_rows["label_order"] = (
        dimension_label_rows.groupby(
            [
                "campaign_scope",
                "dimension_name",
                "platform",
                "buying_type",
                "format_canonical",
            ]
        ).cumcount()
        + 1
    )
    dimension_label_rows["label_key"] = (
        dimension_label_rows["campaign_scope"]
        + "|"
        + dimension_label_rows["dimension_name"]
        + "|"
        + dimension_label_rows["platform"]
        + "|"
        + dimension_label_rows["buying_type"]
        + "|"
        + dimension_label_rows["format_canonical"]
        + "|"
        + dimension_label_rows["label_order"].astype(int).astype(str)
    )

    matrix_rows_exact = (
        creative_rows_exact.groupby(
            [
                "campaign_scope",
                "platform",
                "buying_type",
                "objective_normalized",
                "format_canonical",
            ],
            dropna=False,
        )
        .agg(
            creative_count=("display_name", "count"),
            score_sum=("composite_score", "sum"),
        )
        .reset_index()
        .sort_values(
            [
                "campaign_scope",
                "platform",
                "buying_type",
                "objective_normalized",
                "format_canonical",
            ]
        )
    )
    matrix_rows = (
        _expand_selector_scopes(matrix_rows_exact, include_format_all=False)
        .groupby(
            [
                "campaign_scope",
                "platform",
                "buying_type",
                "objective_normalized",
                "format_canonical",
            ],
            dropna=False,
        )
        .agg(
            creative_count=("creative_count", "sum"),
            score_sum=("score_sum", "sum"),
        )
        .reset_index()
        .sort_values(
            [
                "campaign_scope",
                "platform",
                "buying_type",
                "objective_normalized",
                "format_canonical",
            ]
        )
    )

    campaigns = ["All"] + sorted(
        [
            c
            for c in raw_enriched["campaign_normalized"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
            if c and c != "All"
        ]
    )
    formats = ["All"] + sorted(
        [
            f
            for f in raw_enriched["format_canonical"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
            if f
        ]
    )
    objectives = sorted(
        [
            o
            for o in creative_rows_exact["objective_normalized"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
            if o
        ]
    )

    return {
        "creative_rows": creative_rows,
        "dimension_rows": dimension_rows,
        "dimension_detail_rows": dimension_detail_rows,
        "dimension_label_rows": dimension_label_rows,
        "placement_summary": placement_summary,
        "os_summary": os_summary,
        "asset_summary": asset_summary,
        "matrix_rows": matrix_rows,
        "campaigns": campaigns,
        "formats": formats,
        "objectives": objectives,
    }


def _range_ref(sheet_name: str, col_idx: int, start_row: int, end_row: int) -> str:
    """Excel range reference helper."""
    end_row = max(end_row, start_row)
    col_letter = get_column_letter(col_idx)
    return f"'{sheet_name}'!${col_letter}${start_row}:${col_letter}${end_row}"


def _block_ref(
    sheet_name: str, start_col: int, end_col: int, start_row: int, end_row: int
) -> str:
    """Excel block reference helper."""
    end_row = max(end_row, start_row)
    return (
        f"'{sheet_name}'!${get_column_letter(start_col)}${start_row}:"
        f"${get_column_letter(end_col)}${end_row}"
    )


def _write_helper_sheet(ws, df: pd.DataFrame, hidden: bool = True) -> int:
    """Write a helper dataframe to a worksheet and optionally hide it."""
    ws.freeze_panes = "A2"
    header_fill = PatternFill(fill_type="solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(value, (np.bool_, bool)):
                cell.value = bool(value)
            elif pd.isna(value):
                cell.value = ""
            elif isinstance(value, (np.integer, int)):
                cell.value = int(value)
            elif isinstance(value, (np.floating, float)):
                cell.value = round(float(value), 4)
            else:
                cell.value = str(value)

    for col_idx, _ in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    _apply_header_number_formats(ws, header_row=1)

    if hidden:
        ws.sheet_state = "hidden"

    return max(ws.max_row, 2)


def _format_dashboard_kpi_value(label: str, value) -> str:
    """Match the dashboard's visible KPI text formatting."""
    if value is None or pd.isna(value):
        return ""

    numeric = float(value)
    if label == "Spend":
        return f"€{numeric:,.0f}"
    if label == "Reach":
        return f"{numeric:,.0f}"
    if label == "Frequency":
        return f"{numeric:.2f}x"
    return f"{numeric:.2f}%"


def _format_dashboard_creative_label(record: pd.Series) -> str:
    """Build the visible dashboard creative label with explicit context."""
    creative_name = _clean_dashboard_label(
        record.get("creative_name", record.get("display_name", "")),
        default="Unknown",
    )
    objective = _clean_dashboard_label(
        record.get("objective_normalized", ""), default="Unknown"
    )
    placement = _clean_dashboard_label(
        record.get("placement_canonical", ""), default="Unknown"
    )
    audience = _clean_dashboard_label(
        record.get("audience_segment_label", ""), default="Unknown"
    )
    source_rows = int(pd.to_numeric(record.get("source_rows", 0), errors="coerce") or 0)
    source_label = f"{source_rows:,}" if source_rows > 0 else "0"
    return (
        f"{creative_name}\n"
        f"Objective: {objective} | Placement: {placement} | Audience: {audience} | Source rows: {source_label}"
    )


def _build_validation_formula(values: list[object]) -> str | None:
    """Prefer inline validation lists for cross-app compatibility when safe."""
    cleaned: list[str] = []
    for value in values:
        text = _clean_dashboard_label(value)
        if "," in text:
            return None
        cleaned.append(text.replace('"', '""'))

    formula = '"' + ",".join(cleaned) + '"'
    if len(formula) > 255:
        return None
    return formula


def _build_dashboard_cached_values(
    support: dict, df: pd.DataFrame
) -> dict[str, object]:
    """Precompute the default dashboard view so the workbook opens populated."""
    selectors = {
        "campaign": "All",
        "platform": "All",
        "format": "All",
        "buying": "Paid",
        "top_range": 10,
        "dimension": "Asset Type",
        "kpis": ["VTR", "Completion Rate", "CTR"],
    }
    kpi_columns = {
        "VTR": "vtr_2s",
        "Completion Rate": "completion_rate",
        "CTR": "ctr",
        "Engagement Rate": "engagement_rate",
        "Spend": "spend",
        "Reach": "reach",
        "Frequency": "frequency",
    }

    cache: dict[str, object] = {}
    creative_rows = support["creative_rows"]
    creative_mask = (
        (creative_rows["campaign_scope"] == selectors["campaign"])
        & (creative_rows["platform"] == selectors["platform"])
        & (creative_rows["buying_type"] == selectors["buying"])
        & (creative_rows["format_canonical"] == selectors["format"])
    )
    creative_filtered = creative_rows[creative_mask].copy()

    cache["B11"] = int(
        (creative_mask & ~creative_rows["low_confidence"].fillna(False)).sum()
    )
    if "low_confidence" in df.columns:
        cache["D11"] = int((~df["low_confidence"].fillna(False)).sum())
    else:
        cache["D11"] = int(len(df))
    cache["F11"] = int(
        (
            (creative_rows["campaign_scope"] == selectors["campaign"])
            & (creative_rows["platform"] == "TikTok")
            & (creative_rows["buying_type"] == selectors["buying"])
            & (creative_rows["format_canonical"] == selectors["format"])
            & ~creative_rows["low_confidence"].fillna(False)
        ).sum()
    )
    cache["G11"] = int(
        (
            (creative_rows["campaign_scope"] == selectors["campaign"])
            & (creative_rows["platform"] == "Meta")
            & (creative_rows["buying_type"] == selectors["buying"])
            & (creative_rows["format_canonical"] == selectors["format"])
            & ~creative_rows["low_confidence"].fillna(False)
        ).sum()
    )
    cache["E15"] = selectors["kpis"][0]
    cache["F15"] = selectors["kpis"][1]
    cache["G15"] = selectors["kpis"][2]
    cache["E70"] = selectors["kpis"][0]
    cache["F70"] = selectors["kpis"][1]
    cache["G70"] = selectors["kpis"][2]
    cache["B125"] = f"{selectors['dimension']} groups shown below"

    def _fill_ranked_table(
        start_row: int, start_col: int, ranked_df: pd.DataFrame, rank_col: str
    ):
        rank_letter = get_column_letter(start_col)
        creative_letter = get_column_letter(start_col + 1)
        score_letter = get_column_letter(start_col + 2)
        kpi_letters = [get_column_letter(start_col + 3 + idx) for idx in range(3)]
        by_rank = (
            ranked_df.set_index(rank_col, drop=False)
            if not ranked_df.empty
            else pd.DataFrame()
        )

        for idx, kpi_label in enumerate(selectors["kpis"]):
            cache[f"{kpi_letters[idx]}{start_row + 1}"] = kpi_label

        for rank in range(1, 51):
            row_idx = start_row + 1 + rank
            if (
                rank <= selectors["top_range"]
                and not by_rank.empty
                and rank in by_rank.index
            ):
                record = by_rank.loc[rank]
                if isinstance(record, pd.DataFrame):
                    record = record.iloc[0]
                cache[f"{rank_letter}{row_idx}"] = rank
                cache[f"{creative_letter}{row_idx}"] = _format_dashboard_creative_label(
                    record
                )
                cache[f"{score_letter}{row_idx}"] = round(
                    float(record["composite_score"]), 1
                )
                for kpi_label, cell_ref in zip(selectors["kpis"], kpi_letters):
                    cache[f"{cell_ref}{row_idx}"] = _format_dashboard_kpi_value(
                        kpi_label,
                        record.get(kpi_columns[kpi_label]),
                    )
            else:
                cache[f"{rank_letter}{row_idx}"] = ""
                cache[f"{creative_letter}{row_idx}"] = ""
                cache[f"{score_letter}{row_idx}"] = ""
                for cell_ref in kpi_letters:
                    cache[f"{cell_ref}{row_idx}"] = ""

    _fill_ranked_table(14, 2, creative_filtered.sort_values("rank_desc"), "rank_desc")
    _fill_ranked_table(69, 2, creative_filtered.sort_values("rank_asc"), "rank_asc")

    placement_summary = support["placement_summary"]
    placement_filtered = placement_summary[
        (placement_summary["campaign_scope"] == selectors["campaign"])
        & (placement_summary["platform"] == selectors["platform"])
        & (placement_summary["buying_type"] == selectors["buying"])
        & (placement_summary["format_canonical"] == selectors["format"])
    ].sort_values("rank_desc")
    placement_by_rank = (
        placement_filtered.set_index("rank_desc", drop=False)
        if not placement_filtered.empty
        else pd.DataFrame()
    )
    for rank in range(1, 11):
        row_idx = 15 + rank
        if not placement_by_rank.empty and rank in placement_by_rank.index:
            record = placement_by_rank.loc[rank]
            if isinstance(record, pd.DataFrame):
                record = record.iloc[0]
            cache[f"I{row_idx}"] = record["label"]
            cache[f"J{row_idx}"] = int(record["creatives_count"])
            cache[f"K{row_idx}"] = float(record["avg_score"])
            cache[f"L{row_idx}"] = float(record["total_spend"])
            cache[f"M{row_idx}"] = float(record["avg_vtr"])
        else:
            for col in ["I", "J", "K", "L", "M"]:
                cache[f"{col}{row_idx}"] = ""

    def _fill_summary_rows(
        summary_df: pd.DataFrame, label_rows: dict[str, int], start_col: str
    ):
        summary_filtered = summary_df[
            (summary_df["campaign_scope"] == selectors["campaign"])
            & (summary_df["platform"] == selectors["platform"])
            & (summary_df["buying_type"] == selectors["buying"])
            & (summary_df["format_canonical"] == selectors["format"])
        ]
        for label, row_idx in label_rows.items():
            match = summary_filtered[summary_filtered["label"] == label]
            if match.empty:
                cache[f"J{row_idx}"] = 0
                cache[f"K{row_idx}"] = 0
                cache[f"L{row_idx}"] = 0
                cache[f"M{row_idx}"] = 0
                continue
            record = match.iloc[0]
            cache[f"J{row_idx}"] = int(record["creatives_count"])
            cache[f"K{row_idx}"] = float(record["avg_score"])
            cache[f"L{row_idx}"] = float(record["total_spend"])
            cache[f"M{row_idx}"] = float(record["avg_vtr"])

    _fill_summary_rows(support["os_summary"], {"iOS": 30, "Android": 31}, "I")
    _fill_summary_rows(support["asset_summary"], {"Brand": 38, "Creator": 39}, "I")

    matrix_rows = support["matrix_rows"]
    matrix_filtered = matrix_rows[
        (matrix_rows["campaign_scope"] == selectors["campaign"])
        & (matrix_rows["platform"] == selectors["platform"])
        & (matrix_rows["buying_type"] == selectors["buying"])
    ]
    for row_idx, objective in enumerate(support["objectives"][:12], start=46):
        objective_df = matrix_filtered[
            matrix_filtered["objective_normalized"] == objective
        ]
        for col_letter, fmt in [("J", "Video"), ("K", "Motion"), ("L", "Static")]:
            fmt_df = objective_df[objective_df["format_canonical"] == fmt]
            if selectors["format"] != "All" and selectors["format"] != fmt:
                cache[f"{col_letter}{row_idx}"] = ""
            elif fmt_df.empty or float(fmt_df["creative_count"].sum()) == 0:
                cache[f"{col_letter}{row_idx}"] = ""
            else:
                cache[f"{col_letter}{row_idx}"] = round(
                    float(fmt_df["score_sum"].sum())
                    / float(fmt_df["creative_count"].sum()),
                    1,
                )
        cache[f"M{row_idx}"] = (
            int(objective_df["creative_count"].sum()) if not objective_df.empty else 0
        )

    if selectors["dimension"] == "Asset Type":
        group_labels = ["Brand", "Creator", "", ""]
    elif selectors["dimension"] == "OS":
        group_labels = ["iOS", "Android", "", ""]
    elif selectors["dimension"] == "Device":
        group_labels = ["Mobile", "Desktop", "Other", ""]
    else:
        label_rows = support["dimension_label_rows"]
        label_filtered = label_rows[
            (label_rows["campaign_scope"] == selectors["campaign"])
            & (label_rows["dimension_name"] == selectors["dimension"])
            & (label_rows["platform"] == selectors["platform"])
            & (label_rows["buying_type"] == selectors["buying"])
            & (label_rows["format_canonical"] == selectors["format"])
        ].sort_values("label_order")
        group_labels = label_filtered["dimension_value"].tolist()[:4]
        group_labels.extend([""] * (4 - len(group_labels)))

    for cell_ref, label in zip(["B127", "I127", "B182", "I182"], group_labels):
        cache[cell_ref] = label

    dimension_rows = support["dimension_detail_rows"]
    dimension_filtered = dimension_rows[
        (dimension_rows["campaign_scope"] == selectors["campaign"])
        & (dimension_rows["dimension_name"] == selectors["dimension"])
        & (dimension_rows["platform"] == selectors["platform"])
        & (dimension_rows["buying_type"] == selectors["buying"])
        & (dimension_rows["format_canonical"] == selectors["format"])
    ]

    for start_row, start_col, label in [
        (127, 2, group_labels[0]),
        (127, 9, group_labels[1]),
        (182, 2, group_labels[2]),
        (182, 9, group_labels[3]),
    ]:
        if label:
            group_df = dimension_filtered[
                dimension_filtered["dimension_value"] == label
            ].sort_values("rank_desc")
        else:
            group_df = dimension_filtered.iloc[0:0]
        _fill_ranked_table(start_row, start_col, group_df, "rank_desc")

    return cache


def _inject_formula_cache_values(
    output_path: str, sheet_name: str, cached_values: dict[str, object]
):
    """Inject cached results into formula cells so Excel opens with populated values."""
    if not cached_values:
        return

    with zipfile.ZipFile(output_path, "r") as zin:
        workbook_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))

        rel_map = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root.findall(f"{{{PKG_REL_NS}}}Relationship")
        }
        sheet_xml_path = None
        for sheet in workbook_root.findall(f".//{{{XL_MAIN_NS}}}sheet"):
            if sheet.attrib.get("name") != sheet_name:
                continue
            rel_id = sheet.attrib.get(f"{{{DOC_REL_NS}}}id")
            target = rel_map.get(rel_id, "")
            target = target.lstrip("/")
            if not target.startswith("xl/"):
                target = f"xl/{target}"
            sheet_xml_path = target
            break

        if not sheet_xml_path:
            return

        sheet_root = ET.fromstring(zin.read(sheet_xml_path))
        changed = False
        for cell in sheet_root.findall(f".//{{{XL_MAIN_NS}}}c"):
            coord = cell.attrib.get("r")
            if coord not in cached_values or cell.find(f"{{{XL_MAIN_NS}}}f") is None:
                continue

            value = cached_values[coord]
            value_nodes = cell.findall(f"{{{XL_MAIN_NS}}}v")
            for node in value_nodes:
                cell.remove(node)

            value_node = ET.Element(f"{{{XL_MAIN_NS}}}v")
            if value == "" or value is None:
                cell.attrib.pop("t", None)
            elif isinstance(value, (bool, np.bool_)):
                cell.attrib["t"] = "b"
                value_node.text = "1" if bool(value) else "0"
            elif isinstance(
                value, (int, float, np.integer, np.floating)
            ) and not pd.isna(value):
                cell.attrib.pop("t", None)
                value_node.text = str(value)
            else:
                cell.attrib["t"] = "str"
                value_node.text = str(value)

            cell.append(value_node)
            changed = True

        if not changed:
            return

        temp_path = f"{output_path}.ziptmp"
        with zipfile.ZipFile(temp_path, "w") as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename == sheet_xml_path:
                    data = ET.tostring(
                        sheet_root, encoding="utf-8", xml_declaration=True
                    )
                zout.writestr(info, data)

    os.replace(temp_path, output_path)


def _strip_future_function_prefixes(output_path: str):
    """Strip _xlfn._xlws. and _xlfn. prefixes from formula cells in all sheets.

    xlsxwriter auto-prefixes "future functions" (FILTER, SORT, UNIQUE, etc.)
    with ``_xlfn._xlws.`` in the xlsx XML.  Excel 365 understands this prefix,
    but Google Sheets does not — formulas appear blank.  This function patches
    the xlsx XML in-place so the formulas use the bare function names.
    """
    with zipfile.ZipFile(output_path, "r") as zin:
        sheet_files = [
            info.filename
            for info in zin.infolist()
            if info.filename.startswith("xl/worksheets/sheet")
            and info.filename.endswith(".xml")
        ]
        if not sheet_files:
            return

        patched: dict[str, bytes] = {}
        for sheet_file in sheet_files:
            raw = zin.read(sheet_file)
            sheet_root = ET.fromstring(raw)
            changed = False
            for formula_el in sheet_root.findall(f".//{{{XL_MAIN_NS}}}f"):
                if formula_el.text and "_xlfn." in formula_el.text:
                    formula_el.text = formula_el.text.replace(
                        "_xlfn._xlws.", ""
                    ).replace("_xlfn.", "")
                    changed = True
            if changed:
                patched[sheet_file] = ET.tostring(
                    sheet_root, encoding="utf-8", xml_declaration=True
                )

        if not patched:
            return

        temp_path = f"{output_path}.ziptmp"
        with zipfile.ZipFile(temp_path, "w") as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename in patched:
                    data = patched[info.filename]
                zout.writestr(info, data)

    os.replace(temp_path, output_path)


def _write_dashboard_table(
    ws,
    start_row: int,
    start_col: int,
    title: str,
    source_block: str,
    key_range: str,
    key_expr_template: str,
    top_range_cell: str,
    kpi_header_cells: list[str],
    score_fill: str,
    title_formula: str | None = None,
    merge_title: bool = True,
):
    """Write a visible ranked table backed by helper-sheet keys."""
    title_fill = PatternFill(fill_type="solid", fgColor=score_fill)
    header_fill = PatternFill(fill_type="solid", fgColor="E9EEF6")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    if merge_title:
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=start_row,
            end_column=start_col + 5,
        )
    for col_idx in range(start_col, start_col + 6):
        title_cell = ws.cell(start_row, col_idx)
        title_cell.fill = title_fill
        title_cell.border = border
    cell = ws.cell(start_row, start_col)
    cell.value = title_formula or title
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.alignment = Alignment(horizontal="left")

    headers = ["Rank", "Creative", "Score"]
    for offset, header in enumerate(headers, start=start_col):
        h_cell = ws.cell(start_row + 1, offset, header)
        h_cell.fill = header_fill
        h_cell.font = Font(bold=True)
        h_cell.border = border

    for idx, header_cell in enumerate(kpi_header_cells, start=start_col + 3):
        h_cell = ws.cell(start_row + 1, idx, f"={header_cell}")
        h_cell.fill = header_fill
        h_cell.font = Font(bold=True)
        h_cell.border = border

    rank_col_letter = get_column_letter(start_col)
    kpi_lookup = "INDEX('_DashboardLists'!$I$2:$I$8,MATCH({selector},'_DashboardLists'!$H$2:$H$8,0))"
    for row_idx in range(start_row + 2, start_row + 52):
        rank_row_range = (
            f"${rank_col_letter}${start_row + 2}:{rank_col_letter}{row_idx}"
        )
        rank_expr = f"ROWS({rank_row_range})"
        key_expr = key_expr_template.format(rank=rank_expr)
        match_formula = f"MATCH({key_expr},{key_range},0)"
        ws.row_dimensions[row_idx].height = 30

        rank_cell = ws.cell(row_idx, start_col)
        rank_cell.value = f'=IF(AND({rank_expr}<=VALUE({top_range_cell}),COUNTIF({key_range},{key_expr})>0),{rank_expr},"")'
        rank_cell.border = border

        creative_cell = ws.cell(row_idx, start_col + 1)
        creative_cell.value = (
            f'=IF({rank_col_letter}{row_idx}="","",'
            f"INDEX({source_block},{match_formula},1)&CHAR(10)&"
            f'"Objective: "&INDEX({source_block},{match_formula},2)&'
            f'" | Placement: "&INDEX({source_block},{match_formula},3)&'
            f'" | Audience: "&INDEX({source_block},{match_formula},4)&'
            f'" | Source rows: "&INDEX({source_block},{match_formula},5))'
        )
        creative_cell.border = border
        creative_cell.alignment = Alignment(wrap_text=True, vertical="top")

        score_cell = ws.cell(row_idx, start_col + 2)
        score_cell.value = f'=IF({rank_col_letter}{row_idx}="","",ROUND(INDEX({source_block},{match_formula},6),1))'
        score_cell.number_format = "0.0"
        score_cell.border = border

        for col_idx, selector_cell in zip(
            range(start_col + 3, start_col + 6), kpi_header_cells
        ):
            data_cell = ws.cell(row_idx, col_idx)
            index_formula = kpi_lookup.format(selector=selector_cell)
            raw_value = f"INDEX({source_block},{match_formula},{index_formula})"
            data_cell.value = (
                f'=IF({rank_col_letter}{row_idx}="","",'
                f'IF({selector_cell}="Spend","€"&TEXT({raw_value},"#,##0"),'
                f'IF({selector_cell}="Reach",TEXT({raw_value},"#,##0"),'
                f'IF({selector_cell}="Frequency",TEXT({raw_value},"0.00")&"x",'
                f'TEXT({raw_value},"0.00")&"%"))))'
            )
            data_cell.border = border


def _rebuild_dashboard_with_openpyxl(
    output_path: str,
    df: pd.DataFrame,
    df_raw: "pd.DataFrame | None" = None,
    brand: str = "",
    target: str = "excel",
):
    """Replace the static dashboard with a formula-driven dashboard."""
    support = _build_dashboard_support_data(df, df_raw)
    wb = load_workbook(output_path)

    # Remove old dashboard structures and stale named ranges from the xlsxwriter pass.
    # Note: "Top by" tabs are self-contained (helper data in hidden columns) and
    # don't depend on _DashboardData or _Data* named ranges.
    for sheet_name in [
        "Dashboard",
        "_DashboardLists",
        "_DashboardCreatives",
        "_DashboardDimensions",
        "_DashboardDimensionDetail",
        "_DashboardDimensionLabels",
        "_DashboardPlacementSummary",
        "_DashboardOSSummary",
        "_DashboardAssetSummary",
        "_DashboardMatrix",
    ]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    # Preserve _Data* named ranges (may still be used by other sheets)
    for name in list(wb.defined_names.keys()):
        if not name.startswith("_Data"):
            wb.defined_names.pop(name, None)

    dashboard = wb.create_sheet("Dashboard", 0)
    lists_ws = wb.create_sheet("_DashboardLists", 1)
    creative_ws = wb.create_sheet("_DashboardCreatives", 2)
    dimension_ws = wb.create_sheet("_DashboardDimensions", 3)
    dimension_detail_ws = wb.create_sheet("_DashboardDimensionDetail", 4)
    label_ws = wb.create_sheet("_DashboardDimensionLabels", 5)
    placement_ws = wb.create_sheet("_DashboardPlacementSummary", 6)
    os_ws = wb.create_sheet("_DashboardOSSummary", 7)
    asset_ws = wb.create_sheet("_DashboardAssetSummary", 8)
    matrix_ws = wb.create_sheet("_DashboardMatrix", 9)

    # Helper sheets.
    campaigns = support["campaigns"]
    formats = support["formats"]
    objectives = support["objectives"]

    lists_ws["A1"] = "campaigns"
    for idx, value in enumerate(campaigns, start=2):
        lists_ws[f"A{idx}"] = value
    for idx, value in enumerate(["All", "TikTok", "Meta"], start=2):
        lists_ws[f"B{idx}"] = value
    for idx, value in enumerate(formats, start=2):
        lists_ws[f"C{idx}"] = value
    for idx, value in enumerate(["Paid", "Boosting", "Both"], start=2):
        lists_ws[f"D{idx}"] = value
    for idx, value in enumerate(["5", "10", "25", "50"], start=2):
        lists_ws[f"E{idx}"] = value
    for idx, value in enumerate(
        ["Asset Type", "Placement", "OS", "Device", "Objective", "Audience Segment"],
        start=2,
    ):
        lists_ws[f"F{idx}"] = value
    for idx, value in enumerate(
        [
            "VTR",
            "Completion Rate",
            "CTR",
            "Engagement Rate",
            "Spend",
            "Reach",
            "Frequency",
        ],
        start=2,
    ):
        lists_ws[f"G{idx}"] = value

    lists_ws["H1"] = "kpi_label"
    lists_ws["I1"] = "creative_col_index"
    kpi_map = [
        ("VTR", 7),
        ("Completion Rate", 8),
        ("CTR", 9),
        ("Engagement Rate", 10),
        ("Spend", 11),
        ("Reach", 12),
        ("Frequency", 13),
    ]
    for idx, (label, col_index) in enumerate(kpi_map, start=2):
        lists_ws[f"H{idx}"] = label
        lists_ws[f"I{idx}"] = col_index

    lists_ws["K1"] = "objective"
    for idx, value in enumerate(objectives, start=2):
        lists_ws[f"K{idx}"] = value
    lists_ws["L1"] = "matrix_formats"
    for idx, value in enumerate(["Video", "Motion", "Static"], start=2):
        lists_ws[f"L{idx}"] = value
    lists_ws.sheet_state = "hidden"

    creative_end = _write_helper_sheet(creative_ws, support["creative_rows"])
    dimension_end = _write_helper_sheet(dimension_ws, support["dimension_rows"])
    dimension_detail_end = _write_helper_sheet(
        dimension_detail_ws, support["dimension_detail_rows"]
    )
    label_end = _write_helper_sheet(label_ws, support["dimension_label_rows"])
    placement_end = _write_helper_sheet(placement_ws, support["placement_summary"])
    os_end = _write_helper_sheet(os_ws, support["os_summary"])
    asset_end = _write_helper_sheet(asset_ws, support["asset_summary"])
    matrix_end = _write_helper_sheet(matrix_ws, support["matrix_rows"])

    # Range references used in formulas.
    creative_block = _block_ref("_DashboardCreatives", 5, 17, 2, creative_end)
    creative_campaign = _range_ref("_DashboardCreatives", 1, 2, creative_end)
    creative_platform = _range_ref("_DashboardCreatives", 2, 2, creative_end)
    creative_buying = _range_ref("_DashboardCreatives", 3, 2, creative_end)
    creative_format = _range_ref("_DashboardCreatives", 4, 2, creative_end)
    creative_low_conf = _range_ref("_DashboardCreatives", 20, 2, creative_end)
    creative_key_desc = _range_ref("_DashboardCreatives", 24, 2, creative_end)
    creative_key_asc = _range_ref("_DashboardCreatives", 25, 2, creative_end)

    dim_block = _block_ref("_DashboardDimensionDetail", 7, 19, 2, dimension_detail_end)
    dim_campaign = _range_ref("_DashboardDimensionDetail", 1, 2, dimension_detail_end)
    dim_name = _range_ref("_DashboardDimensionDetail", 2, 2, dimension_detail_end)
    dim_value = _range_ref("_DashboardDimensionDetail", 3, 2, dimension_detail_end)
    dim_platform = _range_ref("_DashboardDimensionDetail", 4, 2, dimension_detail_end)
    dim_buying = _range_ref("_DashboardDimensionDetail", 5, 2, dimension_detail_end)
    dim_format = _range_ref("_DashboardDimensionDetail", 6, 2, dimension_detail_end)
    dim_low_conf = _range_ref("_DashboardDimensionDetail", 22, 2, dimension_detail_end)
    dim_group_key = _range_ref("_DashboardDimensionDetail", 24, 2, dimension_detail_end)

    placement_block = _block_ref("_DashboardPlacementSummary", 5, 9, 2, placement_end)
    placement_campaign = _range_ref("_DashboardPlacementSummary", 1, 2, placement_end)
    placement_platform = _range_ref("_DashboardPlacementSummary", 2, 2, placement_end)
    placement_buying = _range_ref("_DashboardPlacementSummary", 3, 2, placement_end)
    placement_format = _range_ref("_DashboardPlacementSummary", 4, 2, placement_end)
    placement_key = _range_ref("_DashboardPlacementSummary", 11, 2, placement_end)

    os_campaign = _range_ref("_DashboardOSSummary", 1, 2, os_end)
    os_platform = _range_ref("_DashboardOSSummary", 2, 2, os_end)
    os_buying = _range_ref("_DashboardOSSummary", 3, 2, os_end)
    os_format = _range_ref("_DashboardOSSummary", 4, 2, os_end)
    os_label = _range_ref("_DashboardOSSummary", 5, 2, os_end)
    os_creatives = _range_ref("_DashboardOSSummary", 6, 2, os_end)
    os_score = _range_ref("_DashboardOSSummary", 7, 2, os_end)
    os_spend = _range_ref("_DashboardOSSummary", 8, 2, os_end)
    os_vtr = _range_ref("_DashboardOSSummary", 9, 2, os_end)

    asset_campaign = _range_ref("_DashboardAssetSummary", 1, 2, asset_end)
    asset_platform = _range_ref("_DashboardAssetSummary", 2, 2, asset_end)
    asset_buying = _range_ref("_DashboardAssetSummary", 3, 2, asset_end)
    asset_format = _range_ref("_DashboardAssetSummary", 4, 2, asset_end)
    asset_label = _range_ref("_DashboardAssetSummary", 5, 2, asset_end)
    asset_creatives = _range_ref("_DashboardAssetSummary", 6, 2, asset_end)
    asset_score = _range_ref("_DashboardAssetSummary", 7, 2, asset_end)
    asset_spend = _range_ref("_DashboardAssetSummary", 8, 2, asset_end)
    asset_vtr = _range_ref("_DashboardAssetSummary", 9, 2, asset_end)

    label_campaign = _range_ref("_DashboardDimensionLabels", 1, 2, label_end)
    label_dimension = _range_ref("_DashboardDimensionLabels", 2, 2, label_end)
    label_value = _range_ref("_DashboardDimensionLabels", 3, 2, label_end)
    label_platform = _range_ref("_DashboardDimensionLabels", 4, 2, label_end)
    label_buying = _range_ref("_DashboardDimensionLabels", 5, 2, label_end)
    label_format = _range_ref("_DashboardDimensionLabels", 6, 2, label_end)
    label_key = _range_ref("_DashboardDimensionLabels", 8, 2, label_end)

    matrix_campaign = _range_ref("_DashboardMatrix", 1, 2, matrix_end)
    matrix_platform = _range_ref("_DashboardMatrix", 2, 2, matrix_end)
    matrix_buying = _range_ref("_DashboardMatrix", 3, 2, matrix_end)
    matrix_objective = _range_ref("_DashboardMatrix", 4, 2, matrix_end)
    matrix_format = _range_ref("_DashboardMatrix", 5, 2, matrix_end)
    matrix_count = _range_ref("_DashboardMatrix", 6, 2, matrix_end)
    matrix_score_sum = _range_ref("_DashboardMatrix", 7, 2, matrix_end)

    sel_campaign = "$C$6"

    # Dashboard layout and styles.
    dashboard.sheet_view.showGridLines = False
    dashboard.freeze_panes = "B11"
    dashboard.sheet_properties.tabColor = "1A1A2E"
    dashboard.column_dimensions["A"].width = 3
    dashboard.column_dimensions["B"].width = 8
    dashboard.column_dimensions["C"].width = 42
    dashboard.column_dimensions["D"].width = 10
    for col in ["E", "F", "G"]:
        dashboard.column_dimensions[col].width = 13
    dashboard.column_dimensions["H"].width = 3
    dashboard.column_dimensions["I"].width = 14
    dashboard.column_dimensions["J"].width = 40
    dashboard.column_dimensions["K"].width = 10
    for col in ["L", "M", "N"]:
        dashboard.column_dimensions[col].width = 14
    for col in ["O", "P"]:
        dashboard.column_dimensions[col].hidden = True
    dashboard.row_dimensions[9].height = 32

    title_fill = PatternFill(fill_type="solid", fgColor="1A1A2E")
    section_fill = PatternFill(fill_type="solid", fgColor="2C3E50")
    pale_fill = PatternFill(fill_type="solid", fgColor="F8F9FA")
    green_fill = PatternFill(fill_type="solid", fgColor="E8F5E9")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    dashboard.merge_cells("B2:G2")
    dashboard["B2"] = f"WPP Scout Dashboard{f' — {brand}' if brand else ''}"
    dashboard["B2"].fill = title_fill
    dashboard["B2"].font = Font(color="FFFFFF", bold=True, size=18)
    dashboard["B2"].alignment = Alignment(horizontal="left")

    dashboard.merge_cells("B3:G3")
    dashboard["B3"] = (
        "Selectors below actively drive the visible analysis."
        if target == "google_sheets"
        else "Selectors below actively drive the visible analysis. Excel 365 / Excel 2021 required."
    )
    dashboard["B3"].font = Font(italic=True, color="666666")

    dashboard.merge_cells("B5:G5")
    dashboard["B5"] = "Filter Controls"
    dashboard["B5"].fill = section_fill
    dashboard["B5"].font = Font(color="FFFFFF", bold=True)

    dashboard.merge_cells("B9:G10")
    dashboard["B9"] = (
        "Top and Bottom tables show creative-level rollups under the current filters. "
        "The second line under each creative shows Objective, Placement, Audience, and Source rows from the raw export."
    )
    dashboard["B9"].font = Font(italic=True, color="666666")
    dashboard["B9"].alignment = Alignment(wrap_text=True)

    dashboard.merge_cells("I5:N5")
    dashboard["I5"] = "How to Read This Dashboard"
    dashboard["I5"].fill = PatternFill(fill_type="solid", fgColor="5B6C7D")
    dashboard["I5"].font = Font(color="FFFFFF", bold=True, size=12)

    dashboard.merge_cells("I6:N10")
    dashboard["I6"] = (
        "1. Each row is one creative within the selected Campaign, Platform, Buying Type, Format, and Objective.\n\n"
        "2. The same creative can appear more than once if it ran against different objectives.\n\n"
        "3. KPI values are weighted rollups across all matching raw ad lines.\n\n"
        "4. Placement labels are platform-aware, for example Meta Feed vs TikTok In Feed.\n\n"
        "5. The second line under each creative name shows Objective, Placement, Audience, and Source rows in the raw export.\n\n"
        "6. The grouped section at the bottom shows raw-row level tactical splits for the selected dimension.\n\n"
        "7. Scores should be compared within the current Platform and Buying Type view, not across Paid vs Boosting."
    )
    dashboard["I6"].fill = pale_fill
    dashboard["I6"].border = border
    dashboard["I6"].alignment = Alignment(wrap_text=True, vertical="top")

    controls = {
        "B6": "Campaign:",
        "D6": "Platform:",
        "F6": "Format:",
        "B7": "Top Range:",
        "D7": "Dimension:",
        "F7": "Buying Type:",
        "B8": "KPI Slot 1:",
        "D8": "KPI Slot 2:",
        "F8": "KPI Slot 3:",
    }
    for ref, label in controls.items():
        dashboard[ref] = label
        dashboard[ref].font = Font(bold=True, color="555555")

    defaults = {
        "C6": "All",
        "E6": "All",
        "G6": "All",
        "C7": "10",
        "E7": "Asset Type",
        "G7": "Paid",
        "C8": "VTR",
        "E8": "Completion Rate",
        "G8": "CTR",
    }
    for ref, value in defaults.items():
        dashboard[ref] = value
        dashboard[ref].fill = green_fill
        dashboard[ref].border = border
        dashboard[ref].font = Font(bold=True)

    validation_specs = [
        (
            "C6",
            campaigns,
            "DashboardCampaigns",
            f"'_DashboardLists'!$A$2:$A${len(campaigns) + 1}",
        ),
        (
            "E6",
            ["All", "TikTok", "Meta"],
            "DashboardPlatforms",
            "'_DashboardLists'!$B$2:$B$4",
        ),
        (
            "G6",
            formats,
            "DashboardFormats",
            f"'_DashboardLists'!$C$2:$C${len(formats) + 1}",
        ),
        (
            "C7",
            ["5", "10", "25", "50"],
            "DashboardTopRanges",
            "'_DashboardLists'!$E$2:$E$5",
        ),
        (
            "E7",
            [
                "Asset Type",
                "Placement",
                "OS",
                "Device",
                "Objective",
                "Audience Segment",
            ],
            "DashboardDimensions",
            "'_DashboardLists'!$F$2:$F$7",
        ),
        (
            "G7",
            ["Paid", "Boosting", "Both"],
            "DashboardBuyingTypes",
            "'_DashboardLists'!$D$2:$D$4",
        ),
        (
            "C8",
            [
                "VTR",
                "Completion Rate",
                "CTR",
                "Engagement Rate",
                "Spend",
                "Reach",
                "Frequency",
            ],
            "DashboardKpis",
            "'_DashboardLists'!$G$2:$G$8",
        ),
        (
            "E8",
            [
                "VTR",
                "Completion Rate",
                "CTR",
                "Engagement Rate",
                "Spend",
                "Reach",
                "Frequency",
            ],
            "DashboardKpis2",
            "'_DashboardLists'!$G$2:$G$8",
        ),
        (
            "G8",
            [
                "VTR",
                "Completion Rate",
                "CTR",
                "Engagement Rate",
                "Spend",
                "Reach",
                "Frequency",
            ],
            "DashboardKpis3",
            "'_DashboardLists'!$G$2:$G$8",
        ),
    ]
    for cell_ref, values, defined_name, fallback_ref in validation_specs:
        formula = _build_validation_formula(values)
        if formula is None:
            wb.defined_names.add(DefinedName(defined_name, attr_text=fallback_ref))
            formula = f"={defined_name}"
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        dashboard.add_data_validation(dv)
        dv.add(dashboard[cell_ref])

    dashboard["B11"] = (
        f"=SUMPRODUCT(({creative_campaign}={sel_campaign})*"
        f"({creative_platform}=$E$6)*"
        f"({creative_format}=$G$6)*"
        f"({creative_buying}=$G$7)*"
        f"({creative_low_conf}=FALSE))"
    )
    dashboard["B12"] = "Filtered Count"
    dashboard["D11"] = (
        int((~df["low_confidence"].fillna(False)).sum())
        if "low_confidence" in df.columns
        else int(len(df))
    )
    dashboard["D12"] = "Total High-Confidence"
    dashboard["F11"] = (
        f'=SUMPRODUCT(({creative_campaign}={sel_campaign})*({creative_platform}="TikTok")*'
        f"({creative_format}=$G$6)*"
        f"({creative_buying}=$G$7)*({creative_low_conf}=FALSE))"
    )
    dashboard["F12"] = "TikTok"
    dashboard["G11"] = (
        f'=SUMPRODUCT(({creative_campaign}={sel_campaign})*({creative_platform}="Meta")*'
        f"({creative_format}=$G$6)*"
        f"({creative_buying}=$G$7)*({creative_low_conf}=FALSE))"
    )
    dashboard["G12"] = "Meta"
    for ref in ["B11", "D11", "F11", "G11"]:
        dashboard[ref].fill = pale_fill
        dashboard[ref].font = Font(bold=True, size=16)
        dashboard[ref].alignment = Alignment(horizontal="center")
        dashboard[ref].border = border
        dashboard[ref].number_format = _dashboard_number_format("count")
    for ref in ["B12", "D12", "F12", "G12"]:
        dashboard[ref].alignment = Alignment(horizontal="center")
        dashboard[ref].font = Font(color="777777", size=9)
        dashboard[ref].border = border

    top_key_expr = '$C$6&"|"&$E$6&"|"&$G$7&"|"&$G$6&"|"&{rank}'
    _write_dashboard_table(
        dashboard,
        14,
        2,
        "Top Performers",
        creative_block,
        creative_key_desc,
        top_key_expr,
        "$C$7",
        ["$C$8", "$E$8", "$G$8"],
        "2C3E50",
    )
    _write_dashboard_table(
        dashboard,
        69,
        2,
        "Bottom Performers",
        creative_block,
        creative_key_asc,
        top_key_expr,
        "$C$7",
        ["$C$8", "$E$8", "$G$8"],
        "C0392B",
    )

    # Placement summary.
    dashboard.merge_cells("I14:N14")
    dashboard["I14"] = "Placement Summary"
    dashboard["I14"].fill = PatternFill(fill_type="solid", fgColor="16A085")
    dashboard["I14"].font = Font(color="FFFFFF", bold=True, size=12)
    placement_headers = ["Placement", "Creatives", "Avg Score", "Spend", "Avg VTR"]
    for idx, header in enumerate(placement_headers, start=9):
        cell = dashboard.cell(15, idx, header)
        cell.fill = pale_fill
        cell.font = Font(bold=True)
        cell.border = border
    for row_idx in range(16, 26):
        for col_idx in range(9, 14):
            cell = dashboard.cell(row_idx, col_idx)
            col_offset = col_idx - 8
            rank_expr = f"ROWS($I$16:I{row_idx})"
            key_expr = f'$C$6&"|"&$E$6&"|"&$G$7&"|"&$G$6&"|"&{rank_expr}'
            match_formula = f"MATCH({key_expr},{placement_key},0)"
            cell.value = f'=IF(COUNTIF({placement_key},{key_expr})=0,"",INDEX({placement_block},{match_formula},{col_offset}))'
            cell.border = border
            if col_idx == 10:
                cell.number_format = _dashboard_number_format("count")
            elif col_idx == 11:
                cell.number_format = _dashboard_number_format("score")
            elif col_idx == 12:
                cell.number_format = _dashboard_number_format("currency")
            elif col_idx == 13:
                cell.number_format = _dashboard_number_format("rate")

    # OS and Brand/Creator overview.
    dashboard.merge_cells("I28:N28")
    dashboard["I28"] = "iOS vs Android Overview"
    dashboard["I28"].fill = PatternFill(fill_type="solid", fgColor="8E44AD")
    dashboard["I28"].font = Font(color="FFFFFF", bold=True, size=12)

    dashboard.merge_cells("I36:N36")
    dashboard["I36"] = "Brand vs Creator Overview"
    dashboard["I36"].fill = PatternFill(fill_type="solid", fgColor="D35400")
    dashboard["I36"].font = Font(color="FFFFFF", bold=True, size=12)

    summary_headers = ["Group", "Creatives", "Avg Score", "Spend", "Avg VTR"]
    for header_row in [29, 37]:
        for idx, header in enumerate(summary_headers, start=9):
            cell = dashboard.cell(header_row, idx, header)
            cell.fill = pale_fill
            cell.font = Font(bold=True)
            cell.border = border

    os_rows = {30: "iOS", 31: "Android", 32: ""}
    for row_idx, label in os_rows.items():
        dashboard.cell(row_idx, 9).value = label
        dashboard.cell(row_idx, 9).border = border
        if label:
            dashboard.cell(
                row_idx, 10
            ).value = f"=SUMIFS({os_creatives},{os_campaign},$C$6,{os_platform},$E$6,{os_buying},$G$7,{os_format},$G$6,{os_label},$I{row_idx})"
            dashboard.cell(
                row_idx, 11
            ).value = f"=SUMIFS({os_score},{os_campaign},$C$6,{os_platform},$E$6,{os_buying},$G$7,{os_format},$G$6,{os_label},$I{row_idx})"
            dashboard.cell(
                row_idx, 12
            ).value = f"=SUMIFS({os_spend},{os_campaign},$C$6,{os_platform},$E$6,{os_buying},$G$7,{os_format},$G$6,{os_label},$I{row_idx})"
            dashboard.cell(
                row_idx, 13
            ).value = f"=SUMIFS({os_vtr},{os_campaign},$C$6,{os_platform},$E$6,{os_buying},$G$7,{os_format},$G$6,{os_label},$I{row_idx})"
        for col_idx in range(9, 14):
            dashboard.cell(row_idx, col_idx).border = border
        dashboard.cell(row_idx, 10).number_format = _dashboard_number_format("count")
        dashboard.cell(row_idx, 11).number_format = _dashboard_number_format("score")
        dashboard.cell(row_idx, 12).number_format = _dashboard_number_format("currency")
        dashboard.cell(row_idx, 13).number_format = _dashboard_number_format("rate")

    asset_rows = {38: "Brand", 39: "Creator", 40: ""}
    for row_idx, label in asset_rows.items():
        dashboard.cell(row_idx, 9).value = label
        dashboard.cell(row_idx, 9).border = border
        if label:
            dashboard.cell(
                row_idx, 10
            ).value = f"=SUMIFS({asset_creatives},{asset_campaign},$C$6,{asset_platform},$E$6,{asset_buying},$G$7,{asset_format},$G$6,{asset_label},$I{row_idx})"
            dashboard.cell(
                row_idx, 11
            ).value = f"=SUMIFS({asset_score},{asset_campaign},$C$6,{asset_platform},$E$6,{asset_buying},$G$7,{asset_format},$G$6,{asset_label},$I{row_idx})"
            dashboard.cell(
                row_idx, 12
            ).value = f"=SUMIFS({asset_spend},{asset_campaign},$C$6,{asset_platform},$E$6,{asset_buying},$G$7,{asset_format},$G$6,{asset_label},$I{row_idx})"
            dashboard.cell(
                row_idx, 13
            ).value = f"=SUMIFS({asset_vtr},{asset_campaign},$C$6,{asset_platform},$E$6,{asset_buying},$G$7,{asset_format},$G$6,{asset_label},$I{row_idx})"
        for col_idx in range(9, 14):
            dashboard.cell(row_idx, col_idx).border = border
        dashboard.cell(row_idx, 10).number_format = _dashboard_number_format("count")
        dashboard.cell(row_idx, 11).number_format = _dashboard_number_format("score")
        dashboard.cell(row_idx, 12).number_format = _dashboard_number_format("currency")
        dashboard.cell(row_idx, 13).number_format = _dashboard_number_format("rate")

    # Objective x Format matrix.
    dashboard.merge_cells("I44:N44")
    dashboard["I44"] = "Objective × Format Matrix"
    dashboard["I44"].fill = PatternFill(fill_type="solid", fgColor="16A085")
    dashboard["I44"].font = Font(color="FFFFFF", bold=True, size=12)
    dashboard["I45"] = "Objective"
    dashboard["I45"].fill = pale_fill
    dashboard["I45"].font = Font(bold=True)
    for idx, fmt in enumerate(["Video", "Motion", "Static"], start=10):
        dashboard.cell(45, idx, fmt).fill = pale_fill
        dashboard.cell(45, idx, fmt).font = Font(bold=True)
        dashboard.cell(45, idx).border = border
    dashboard["M45"] = "Count"
    dashboard["M45"].fill = pale_fill
    dashboard["M45"].font = Font(bold=True)

    for idx, objective in enumerate(objectives[:12], start=46):
        dashboard.cell(idx, 9, objective)
        dashboard.cell(idx, 9).border = border
        for col_idx, fmt in zip([10, 11, 12], ["Video", "Motion", "Static"]):
            numerator = (
                f"SUMPRODUCT(({matrix_campaign}={sel_campaign})*"
                f"({matrix_platform}=$E$6)*"
                f"({matrix_buying}=$G$7)*"
                f"({matrix_objective}=$I{idx})*({matrix_format}={get_column_letter(col_idx)}$45),"
                f"{matrix_score_sum})"
            )
            denominator = (
                f"SUMPRODUCT(({matrix_campaign}={sel_campaign})*"
                f"({matrix_platform}=$E$6)*"
                f"({matrix_buying}=$G$7)*"
                f"({matrix_objective}=$I{idx})*({matrix_format}={get_column_letter(col_idx)}$45),"
                f"{matrix_count})"
            )
            dashboard.cell(idx, col_idx).value = (
                f'=IF(AND($G$6<>"All",{get_column_letter(col_idx)}$45<>$G$6),"",'
                f'IFERROR(ROUND({numerator}/{denominator},1),""))'
            )
            dashboard.cell(idx, col_idx).number_format = "0.0"
            dashboard.cell(idx, col_idx).border = border
        dashboard.cell(idx, 13).value = (
            f"=SUMPRODUCT(({matrix_campaign}={sel_campaign})*"
            f"({matrix_platform}=$E$6)*"
            f"({matrix_buying}=$G$7)*"
            f"({matrix_objective}=$I{idx}),{matrix_count})"
        )
        dashboard.cell(idx, 13).border = border
        dashboard.cell(idx, 13).number_format = _dashboard_number_format("count")

    # Dimension-driven grouped output.
    dashboard.merge_cells("B124:G124")
    dashboard["B124"] = "Dimension-Driven Grouped Output"
    dashboard["B124"].fill = section_fill
    dashboard["B124"].font = Font(color="FFFFFF", bold=True, size=12)
    dashboard["B125"] = '=$E$7&" groups shown below"'
    dashboard["B125"].font = Font(italic=True, color="666666")

    helper_group_formulas = [
        (
            "O127",
            'IF($E$7="Asset Type","Brand",IF($E$7="OS","iOS",IFERROR(INDEX({label_value},MATCH($C$6&"|"&$E$7&"|"&$E$6&"|"&$G$7&"|"&$G$6&"|"&1,{label_key},0)),"")))',
        ),
        (
            "O128",
            'IF($E$7="Asset Type","Creator",IF($E$7="OS","Android",IFERROR(INDEX({label_value},MATCH($C$6&"|"&$E$7&"|"&$E$6&"|"&$G$7&"|"&$G$6&"|"&2,{label_key},0)),"")))',
        ),
        (
            "O129",
            'IF(OR($E$7="Asset Type",$E$7="OS"),"",IFERROR(INDEX({label_value},MATCH($C$6&"|"&$E$7&"|"&$E$6&"|"&$G$7&"|"&$G$6&"|"&3,{label_key},0)),""))',
        ),
        (
            "O130",
            'IF(OR($E$7="Asset Type",$E$7="OS"),"",IFERROR(INDEX({label_value},MATCH($C$6&"|"&$E$7&"|"&$E$6&"|"&$G$7&"|"&$G$6&"|"&4,{label_key},0)),""))',
        ),
    ]
    for cell_ref, formula_template in helper_group_formulas:
        dashboard[cell_ref] = "=" + formula_template.format(
            label_value=label_value, label_key=label_key
        )

    group_configs = [
        (
            127,
            2,
            "$O$127",
            "=$O$127",
        ),
        (
            127,
            9,
            "$O$128",
            "=$O$128",
        ),
        (
            182,
            2,
            "$O$129",
            "=$O$129",
        ),
        (
            182,
            9,
            "$O$130",
            "=$O$130",
        ),
    ]
    for start_row, start_col, label_ref, title_formula in group_configs:
        _write_dashboard_table(
            dashboard,
            start_row,
            start_col,
            "",
            dim_block,
            dim_group_key,
            f'$C$6&"|"&$E$7&"|"&$E$6&"|"&$G$7&"|"&$G$6&"|"&{label_ref}&"|"&{{rank}}',
            "$C$7",
            ["$C$8", "$E$8", "$G$8"],
            "34495E",
            title_formula=title_formula,
            merge_title=(target != "google_sheets"),
        )

    for sheet in wb.worksheets:
        if sheet.title == "Dashboard":
            continue
        if sheet.title.startswith("_Dashboard"):
            continue
        if sheet.title == "How Scoring Works":
            continue
        header_row = 2 if sheet.title == "Summary - Top Performers" else 1
        _apply_header_number_formats(sheet, header_row=header_row)

    wb.calculation.calcId = 0
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcCompleted = False
    wb.calculation.calcOnSave = True
    wb.active = wb.sheetnames.index("Dashboard")
    dashboard.sheet_view.tabSelected = True
    wb.save(output_path)
    if target != "google_sheets":
        _inject_formula_cache_values(
            output_path, "Dashboard", _build_dashboard_cached_values(support, df)
        )
    else:
        _strip_future_function_prefixes(output_path)


# =============================================================================
# HELPER FUNCTIONS FOR SHEET CREATION
# =============================================================================


def _write_top_by_dimension_sheet(
    writer,
    workbook,
    df: pd.DataFrame,
    dimension_name: str,
    dimension_col: str,
    header_fmt,
    campaigns=None,
    target: str = "excel",
):
    """Create a flat filterable table showing top performers by a specific dimension.

    Writes a single table with the dimension value and campaign as columns,
    sorted by dimension then score. Users can apply their own filters in
    Excel or Google Sheets.
    """
    sheet_name = f"Top by {dimension_name}"
    ws = workbook.add_worksheet(sheet_name)
    ws.hide_gridlines(2)
    ws.set_tab_color("#3498db")

    if dimension_col not in df.columns:
        ws.write(0, 0, f'Dimension column "{dimension_col}" not found in data')
        return

    # Filter to high-confidence rows
    helper_df = df[~df["low_confidence"]].copy()

    # Explode comma-separated dimension values (e.g. os_target="Android, iOS")
    # so each value gets its own row for proper filtering.
    if (
        dimension_col in helper_df.columns
        and helper_df[dimension_col].astype(str).str.contains(",", na=False).any()
    ):
        helper_df[dimension_col] = (
            helper_df[dimension_col].astype(str).str.split(r",\s*")
        )
        helper_df = helper_df.explode(dimension_col)
        helper_df[dimension_col] = helper_df[dimension_col].str.strip()
        helper_df = helper_df[helper_df[dimension_col].str.len() > 0]

    if "composite_score" in helper_df.columns:
        helper_df = helper_df.sort_values(
            [dimension_col, "composite_score"], ascending=[True, False]
        ).reset_index(drop=True)

    # Flat table columns — dimension and campaign are filterable columns
    display_headers = [
        dimension_name,
        "Campaign",
        "Creative",
        "Platform",
        "Score",
        "Tier",
        "Spend",
        "Reach",
        "VTR",
        "Completion Rate",
    ]
    data_col_map = [
        dimension_col,
        "campaign_normalized",
        "creative_name",
        "platform",
        "composite_score",
        "tier",
        "spend",
        "reach",
        "vtr_2s",
        "completion_rate",
    ]

    # Write header row
    for ci, h in enumerate(display_headers):
        ws.write(0, ci, h, header_fmt)

    # Write data rows
    for row_idx in range(len(helper_df)):
        r = helper_df.iloc[row_idx]
        for ci, col_name in enumerate(data_col_map):
            val = r.get(col_name, "")
            if pd.isna(val):
                val = ""
            elif isinstance(val, (float, np.floating)):
                val = round(val, 2)
            ws.write(row_idx + 1, ci, val)

    # Add autofilter so users can filter by dimension, campaign, etc.
    last_row = len(helper_df)
    ws.autofilter(0, 0, last_row, len(display_headers) - 1)

    # Conditional formatting on Score column
    score_col = display_headers.index("Score")
    score_letter = chr(ord("A") + score_col)
    if last_row > 0:
        ws.conditional_format(
            f"{score_letter}2:{score_letter}{last_row + 1}",
            {
                "type": "3_color_scale",
                "min_color": "#f8d7da",
                "mid_color": "#fff3cd",
                "max_color": "#d4edda",
                "min_type": "num",
                "min_value": 0,
                "mid_type": "num",
                "mid_value": 50,
                "max_type": "num",
                "max_value": 100,
            },
        )

    # Column widths
    ws.set_column(0, 0, 16)  # Dimension
    ws.set_column(1, 1, 28)  # Campaign
    ws.set_column(2, 2, 42)  # Creative
    ws.set_column(3, len(display_headers) - 1, 14)


def _write_rankings_sheet(
    writer, workbook, df: pd.DataFrame, sheet_name: str, header_fmt, wrap_fmt
):
    """Create a full rankings sheet for a platform/cohort."""
    display_cols = [c for c in DISPLAY_COLS if c in df.columns]
    platform_df = df[display_cols].copy()

    platform_df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    for i, col in enumerate(platform_df.columns):
        ws.write(0, i, col, header_fmt)

    ws.set_column("A:A", 42)
    ws.set_column("B:Z", 14)
    last_col_idx = len(platform_df.columns) - 1
    ws.set_column(last_col_idx, last_col_idx, 80, wrap_fmt)
    ws.autofilter(0, 0, len(platform_df), len(platform_df.columns) - 1)

    # Conditional formatting on score
    if "composite_score" in platform_df.columns:
        score_idx = list(platform_df.columns).index("composite_score")
        score_letter = chr(ord("A") + score_idx)
        n_rows = len(platform_df)
        ws.conditional_format(
            f"{score_letter}2:{score_letter}{n_rows + 1}",
            {
                "type": "3_color_scale",
                "min_color": "#f8d7da",
                "mid_color": "#fff3cd",
                "max_color": "#d4edda",
                "min_type": "num",
                "min_value": 0,
                "mid_type": "num",
                "mid_value": 50,
                "max_type": "num",
                "max_value": 100,
            },
        )


def _write_splits_sheet(writer, workbook, df_raw: pd.DataFrame, header_fmt):
    """Raw variant data — all ad lines with split dimensions."""
    cols = [
        "creative_name",
        "platform",
        "format_canonical",
        "placement_canonical",
        "objective_normalized",
        "asset_type_canonical",
        "asset_type_subtype",
        "os_target",
        "audience_segment",
        "composite_score",
        "tier",
        "buying_type",
        "spend",
        "reach",
        "impressions",
        "frequency",
        "vtr_2s",
        "completion_rate",
        "ctr",
        "engagement_rate",
        "attention_proxy_score",
        "attention_inputs_available",
        "score_renormalized",
        "low_confidence",
    ]
    out_cols = [c for c in cols if c in df_raw.columns]
    out = df_raw[out_cols].copy()
    out.to_excel(writer, sheet_name="Splits Analysis", index=False)
    ws = writer.sheets["Splits Analysis"]
    for i, col in enumerate(out.columns):
        ws.write(0, i, col, header_fmt)
    ws.set_column("A:A", 40)
    ws.set_column("B:W", 14)
    ws.autofilter(0, 0, len(out), len(out.columns) - 1)
    ws.set_tab_color("#2980b9")


def _write_os_comparison_sheet(writer, workbook, df_raw: pd.DataFrame, header_fmt):
    """iOS vs Android performance comparison."""
    os_df = df_raw[df_raw["os_target"].isin(["iOS", "Android"])].copy()
    if os_df.empty:
        return

    # Aggregate per creative × platform × os_target
    agg_cols = {
        "spend": ("spend", "sum"),
        "impressions": ("impressions", "sum"),
        "reach": ("reach", "sum"),
        "vtr_2s": ("vtr_2s", "mean"),
        "completion_rate": ("completion_rate", "mean"),
    }
    if "composite_score" in os_df.columns:
        agg_cols["composite_score"] = ("composite_score", "mean")

    grp = os_df.groupby(["creative_name", "platform", "os_target"], as_index=False).agg(
        **agg_cols
    )

    # Pivot for side-by-side
    pivot = grp.pivot_table(
        index=["creative_name", "platform"],
        columns="os_target",
        values=["spend", "reach", "vtr_2s", "completion_rate"]
        + (["composite_score"] if "composite_score" in grp.columns else []),
        aggfunc="first",
    ).reset_index()

    # Flatten MultiIndex columns for Excel export
    if isinstance(pivot.columns, pd.MultiIndex):
        pivot.columns = [f"{v}_{os}" if os else v for v, os in pivot.columns]

    pivot.to_excel(writer, sheet_name="OS Comparison", index=False)
    ws = writer.sheets["OS Comparison"]
    for i, col in enumerate(pivot.columns):
        ws.write(0, i, str(col), header_fmt)
    ws.set_column("A:A", 40)
    ws.set_column("B:Z", 16)
    ws.autofilter(0, 0, len(pivot), len(pivot.columns) - 1)
    ws.set_tab_color("#8e44ad")


def _write_asset_type_sheet(writer, workbook, df_raw: pd.DataFrame, header_fmt):
    """Brand vs Creator performance comparison."""
    asset_col = (
        "asset_type_canonical"
        if "asset_type_canonical" in df_raw.columns
        else "asset_type"
    )
    if asset_col not in df_raw.columns:
        return

    agg_metrics = {
        "composite_score": "mean",
        "vtr_2s": "mean",
        "completion_rate": "mean",
        "spend": "sum",
        "reach": "sum",
        "impressions": "sum",
        "creative_name": "nunique",
    }
    available = {k: v for k, v in agg_metrics.items() if k in df_raw.columns}

    by_type = df_raw.groupby([asset_col, "platform"], as_index=False).agg(available)
    by_type = by_type.rename(columns={"creative_name": "n_creatives"})

    by_type.to_excel(writer, sheet_name="Asset Types", index=False)
    ws = writer.sheets["Asset Types"]
    for i, col in enumerate(by_type.columns):
        ws.write(0, i, col, header_fmt)
    ws.set_column("A:J", 18)
    ws.autofilter(0, 0, len(by_type), len(by_type.columns) - 1)
    ws.set_tab_color("#e67e22")


def _write_cross_platform_sheet(writer, workbook, cross: pd.DataFrame, header_fmt):
    """Creatives running on both TikTok and Meta."""
    metrics = [
        "spend",
        "reach",
        "impressions",
        "vtr_2s",
        "completion_rate",
        "composite_score",
    ]
    available_metrics = [m for m in metrics if m in cross.columns]

    pivot = cross.pivot_table(
        index="creative_name",
        columns="platform",
        values=available_metrics,
        aggfunc="mean",
    ).reset_index()

    pivot.to_excel(writer, sheet_name="Cross-Platform", index=False)
    ws = writer.sheets["Cross-Platform"]
    for i, col in enumerate(pivot.columns):
        ws.write(0, i, str(col), header_fmt)
    ws.set_column("A:A", 40)
    ws.set_column("B:Z", 16)
    ws.autofilter(0, 0, len(pivot), len(pivot.columns) - 1)
    ws.set_tab_color("#27ae60")


def _write_objective_format_matrix(writer, workbook, df: pd.DataFrame, header_fmt):
    """Average score by objective × format cross-tab."""
    if "composite_score" not in df.columns:
        return

    format_col = "format_canonical" if "format_canonical" in df.columns else "format"
    if format_col not in df.columns:
        return

    pivot = (
        df.pivot_table(
            index="objective_normalized"
            if "objective_normalized" in df.columns
            else "objective",
            columns=format_col,
            values="composite_score",
            aggfunc="mean",
        )
        .round(1)
        .reset_index()
    )

    pivot.to_excel(writer, sheet_name="Obj × Format Matrix", index=False)
    ws = writer.sheets["Obj × Format Matrix"]
    for i, col in enumerate(pivot.columns):
        ws.write(0, i, str(col), header_fmt)
    ws.set_column("A:A", 20)
    ws.set_column("B:P", 16)
    ws.set_tab_color("#c0392b")

    # Heat colouring
    n_rows = len(pivot)
    for ci in range(1, len(pivot.columns)):
        col_letter = chr(ord("A") + ci)
        ws.conditional_format(
            f"{col_letter}2:{col_letter}{n_rows + 1}",
            {
                "type": "3_color_scale",
                "min_color": "#f8d7da",
                "mid_color": "#fff3cd",
                "max_color": "#d4edda",
                "min_type": "num",
                "min_value": 0,
                "mid_type": "num",
                "mid_value": 50,
                "max_type": "num",
                "max_value": 100,
            },
        )


def _write_looker_sheet(writer, workbook, df_raw: pd.DataFrame, header_fmt):
    """Flat export for Looker Studio connection."""
    out_cols = [c for c in LOOKER_COLS if c in df_raw.columns]
    out = df_raw[out_cols].copy()

    # Cast categoricals and booleans to string for Looker compatibility
    for bool_col in [
        "tier",
        "low_confidence",
        "cross_platform",
        "attention_inputs_available",
        "score_renormalized",
    ]:
        if bool_col in out.columns:
            out[bool_col] = out[bool_col].astype(str)

    out.to_excel(writer, sheet_name="Looker Export", index=False)
    ws = writer.sheets["Looker Export"]
    for i, col in enumerate(out.columns):
        ws.write(0, i, col, header_fmt)
    ws.set_column("A:B", 40)
    ws.set_column("C:Z", 16)
    ws.autofilter(0, 0, len(out), len(out.columns) - 1)
    ws.set_tab_color("#16a085")


def export_looker_csv(df_raw: pd.DataFrame, output_path: str):
    """Export scored variants as XLSX for Looker Studio."""
    out_cols = [c for c in LOOKER_COLS if c in df_raw.columns]
    out = df_raw[out_cols].copy()

    for bool_col in [
        "tier",
        "low_confidence",
        "cross_platform",
        "attention_inputs_available",
        "score_renormalized",
    ]:
        if bool_col in out.columns:
            out[bool_col] = out[bool_col].astype(str)

    if not output_path.endswith(".xlsx"):
        output_path = output_path.rsplit(".", 1)[0] + ".xlsx"

    with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
        out.to_excel(writer, sheet_name="Looker Data", index=False)
        workbook = writer.book
        ws = writer.sheets["Looker Data"]
        header_fmt = workbook.add_format(
            {
                "bold": True,
                "bg_color": "#1a1a2e",
                "font_color": "white",
                "border": 1,
            }
        )
        for i, col in enumerate(out.columns):
            ws.write(0, i, col, header_fmt)
        ws.set_column("A:B", 40)
        ws.set_column("C:Z", 16)
        ws.autofilter(0, 0, len(out), len(out.columns) - 1)

    print(f"Looker export saved to: {output_path}")
